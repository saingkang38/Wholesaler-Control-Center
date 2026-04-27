import io
import json
import time
from flask import render_template, request, redirect, url_for, flash, jsonify, Response, stream_with_context, current_app, send_file
from flask_login import login_required
from sqlalchemy import func, case
from app.infrastructure import db
from app.store import store_bp
from app.store.models import StoreProduct, NaverStore, SyncLog
from app.master.models import MasterProduct
from app.wholesalers.models import Wholesaler
from app.actions.models import ActionSignal


# 우리 시스템에서 제거된 도매처 prefix (옛 거래처) — 향후 추가 시 이 목록 갱신
DEPRECATED_PREFIXES = ("DOTO_", "ONCH_")


def _active_prefixes() -> list:
    """현재 활성 도매처의 prefix 목록 — wholesalers 테이블에서 동적 조회."""
    rows = Wholesaler.query.filter(
        Wholesaler.is_active.is_(True),
        Wholesaler.prefix.isnot(None),
    ).all()
    return [w.prefix for w in rows if w.prefix]


def _prefix_to_wholesaler_name() -> dict:
    """prefix → 도매처 이름 매핑 (활성·비활성 모두)."""
    return {
        w.prefix: w.name
        for w in Wholesaler.query.filter(Wholesaler.prefix.isnot(None)).all()
    }


def _classify_unmatched(seller_code: str, active_prefixes: list) -> tuple:
    """미매칭 store_product 의 seller_management_code 를 자동 분류.
    Returns: (group, suspected_prefix)
      group ∈ {'empty', 'active', 'deprecated', 'unknown'}
    """
    if not seller_code or not seller_code.strip():
        return "empty", None
    for p in active_prefixes:
        if p and seller_code.startswith(p):
            return "active", p
    for p in DEPRECATED_PREFIXES:
        if seller_code.startswith(p):
            return "deprecated", p
    return "unknown", None


STATUS_LABELS = {
    "WAIT":        "판매대기",
    "SALE":        "판매중",
    "SOLDOUT":     "품절",
    "SUSPENSION":  "판매중지",
    "CLOSE":       "판매종료",
    "PROHIBITION": "판매금지",
}


def _all_stores():
    return NaverStore.query.order_by(NaverStore.store_name).all()


@store_bp.route("/store-overview")
@login_required
def store_overview_page():
    stores = _all_stores()
    store_id = request.args.get("store_id", type=int)
    if not store_id and stores:
        store_id = stores[0].id

    selected_store = NaverStore.query.get(store_id) if store_id else None
    data = []
    unmatched = 0
    totals = {"total": 0, "sale": 0, "soldout": 0, "price_changes": 0}

    signal_counts = {"SUSPEND_NEEDED": 0, "DISCONTINUE_NEEDED": 0,
                     "PRICE_UP_NEEDED": 0, "PRICE_DOWN_POSSIBLE": 0, "RESUME_POSSIBLE": 0,
                     "OPTION_ADD": 0, "OPTION_PRICE_CHANGE": 0, "OPTION_STOCK_CHANGE": 0}

    if store_id:
        rows = db.session.query(
            Wholesaler.id,
            Wholesaler.name,
            func.count(StoreProduct.id).label("total"),
            func.sum(case((StoreProduct.store_status == "SALE", 1), else_=0)).label("sale"),
            func.sum(case((StoreProduct.store_status == "SOLDOUT", 1), else_=0)).label("soldout"),
        ).join(MasterProduct, StoreProduct.master_product_id == MasterProduct.id)\
         .join(Wholesaler, MasterProduct.wholesaler_id == Wholesaler.id)\
         .filter(StoreProduct.naver_store_id == store_id)\
         .group_by(Wholesaler.id, Wholesaler.name)\
         .all()

        price_rows = db.session.query(
            MasterProduct.wholesaler_id,
            func.count(ActionSignal.id).label("cnt")
        ).join(StoreProduct, ActionSignal.store_product_id == StoreProduct.id)\
         .join(MasterProduct, StoreProduct.master_product_id == MasterProduct.id)\
         .filter(StoreProduct.naver_store_id == store_id)\
         .filter(ActionSignal.status == "pending")\
         .filter(ActionSignal.signal_type.in_(["PRICE_UP_NEEDED", "PRICE_DOWN_POSSIBLE"]))\
         .group_by(MasterProduct.wholesaler_id)\
         .all()
        price_map = {r.wholesaler_id: r.cnt for r in price_rows}

        option_rows = db.session.query(
            MasterProduct.wholesaler_id,
            func.count(ActionSignal.id).label("cnt")
        ).join(StoreProduct, ActionSignal.store_product_id == StoreProduct.id)\
         .join(MasterProduct, StoreProduct.master_product_id == MasterProduct.id)\
         .filter(StoreProduct.naver_store_id == store_id)\
         .filter(ActionSignal.status == "pending")\
         .filter(ActionSignal.signal_type.in_(["OPTION_ADD", "OPTION_PRICE_CHANGE", "OPTION_STOCK_CHANGE"]))\
         .group_by(MasterProduct.wholesaler_id)\
         .all()
        option_map = {r.wholesaler_id: r.cnt for r in option_rows}

        totals["option_changes"] = 0
        for row in rows:
            sale = row.sale or 0
            soldout = row.soldout or 0
            price_changes = price_map.get(row.id, 0)
            option_changes = option_map.get(row.id, 0)
            data.append({
                "wholesaler_name": row.name,
                "total": row.total,
                "sale": sale,
                "soldout": soldout,
                "price_changes": price_changes,
                "option_changes": option_changes,
            })
            totals["total"] += row.total
            totals["sale"] += sale
            totals["soldout"] += soldout
            totals["price_changes"] += price_changes
            totals["option_changes"] += option_changes

        unmatched = StoreProduct.query.filter_by(
            naver_store_id=store_id, master_product_id=None
        ).count()

        # 신호 타입별 카운트
        sig_rows = db.session.query(
            ActionSignal.signal_type,
            func.count(ActionSignal.id)
        ).join(StoreProduct, ActionSignal.store_product_id == StoreProduct.id)\
         .filter(StoreProduct.naver_store_id == store_id)\
         .filter(ActionSignal.status == "pending")\
         .group_by(ActionSignal.signal_type).all()
        for sig_type, cnt in sig_rows:
            if sig_type in signal_counts:
                signal_counts[sig_type] = cnt

    return render_template(
        "store_overview.html",
        stores=stores,
        selected_store=selected_store,
        data=data,
        unmatched=unmatched,
        totals=totals,
        signal_counts=signal_counts,
    )


@store_bp.route("/store-overview/sync", methods=["POST"])
@login_required
def store_overview_sync():
    """동기화 백그라운드 시작 — JSON 응답"""
    from app.store import start_sync_background
    store_id = request.json.get("store_id") if request.is_json else request.form.get("store_id", type=int)
    if not store_id:
        return jsonify({"error": "store_id 필요"}), 400
    NaverStore.query.get_or_404(store_id)
    flask_app = current_app._get_current_object()
    started = start_sync_background(int(store_id), flask_app)
    return jsonify({"started": started, "already_running": not started})


@store_bp.route("/store-overview/sync-stream/<int:store_id>")
@login_required
def sync_stream(store_id):
    """SSE — 실시간 동기화 진행 스트림"""
    from app.store import _sync_progress, _sync_lock

    def generate():
        last_idx = 0
        while True:
            with _sync_lock:
                p = _sync_progress.get(store_id)

            if not p:
                yield f"data: {json.dumps({'log': '진행 정보 없음', 'percent': 0, 'done': True})}\n\n"
                break

            logs = p["logs"]
            percent = p["percent"]
            done = p["done"]
            error = p.get("error")

            for msg in logs[last_idx:]:
                yield f"data: {json.dumps({'log': msg, 'percent': percent})}\n\n"
                last_idx += 1

            if done:
                yield f"data: {json.dumps({'done': True, 'percent': percent, 'error': error})}\n\n"
                break

            time.sleep(0.4)

    return Response(
        stream_with_context(generate()),
        mimetype="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


@store_bp.route("/store-overview/rematch-codes", methods=["POST"])
@login_required
def store_rematch_codes():
    from app.store import _rematch_by_codes
    store_id = request.form.get("store_id", type=int)
    raw = request.form.get("codes", "")
    codes = [c.strip() for c in raw.splitlines() if c.strip()]

    if not store_id or not codes:
        flash("스토어와 코드 목록을 입력해주세요.", "error")
        return redirect(url_for("store.store_overview_page", store_id=store_id))

    try:
        stats = _rematch_by_codes(store_id, codes)
        detail = f"매칭 {stats['matched']} / 미매칭 {stats['unmatched']} / DB없음 {stats['not_found']}"
        flash(f"부분 재매칭 완료 — {detail}", "success")
        SyncLog.query.filter_by(naver_store_id=store_id, action="REMATCH").delete()
        db.session.add(SyncLog(naver_store_id=store_id, action="REMATCH", result="success", detail=detail))
        db.session.commit()
    except Exception as e:
        SyncLog.query.filter_by(naver_store_id=store_id, action="REMATCH").delete()
        db.session.add(SyncLog(naver_store_id=store_id, action="REMATCH", result="error", detail=str(e)))
        db.session.commit()
        flash(f"재매칭 실패: {e}", "error")
    return redirect(url_for("store.store_overview_page", store_id=store_id))


@store_bp.route("/store-overview/proposals")
@login_required
def proposals_page():
    from app.store import propose_code_matches
    store_id = request.args.get("store_id", type=int)
    stores = _all_stores()
    if not store_id and stores:
        store_id = stores[0].id
    selected_store = NaverStore.query.get(store_id) if store_id else None
    proposals = propose_code_matches(store_id) if store_id else []
    return render_template(
        "store_proposals.html",
        stores=stores,
        selected_store=selected_store,
        proposals=proposals,
    )


@store_bp.route("/store-overview/propose-matches")
@login_required
def propose_matches():
    from app.store import propose_code_matches
    store_id = request.args.get("store_id", type=int)
    wholesaler_id = request.args.get("wholesaler_id", type=int)
    if not store_id:
        return jsonify({"error": "store_id 필요"}), 400
    results = propose_code_matches(store_id, wholesaler_id)
    return jsonify(results)


@store_bp.route("/store-overview/apply-proposals", methods=["POST"])
@login_required
def apply_proposals():
    from app.store import _rematch_by_codes
    data = request.get_json()
    store_id = data.get("store_id")
    codes = data.get("codes", [])
    if not store_id or not codes:
        return jsonify({"error": "store_id와 codes 필요"}), 400
    try:
        stats = _rematch_by_codes(store_id, codes)
        db.session.add(SyncLog(
            naver_store_id=store_id,
            action="APPLY_PROPOSALS",
            result="success",
            detail=f"매칭 {stats['matched']} / 미매칭 {stats['unmatched']} / DB없음 {stats['not_found']}",
        ))
        db.session.commit()
        return jsonify(stats)
    except Exception as e:
        db.session.add(SyncLog(naver_store_id=store_id, action="APPLY_PROPOSALS", result="error", detail=str(e)))
        db.session.commit()
        return jsonify({"error": str(e)}), 500


@store_bp.route("/store-overview/push-codes", methods=["POST"])
@login_required
def push_codes():
    from app.store import push_seller_management_codes
    data = request.get_json()
    store_id = data.get("store_id")
    pairs = data.get("pairs", [])
    if not store_id or not pairs:
        return jsonify({"error": "store_id와 pairs 필요"}), 400
    naver_store = NaverStore.query.get_or_404(store_id)
    try:
        result = push_seller_management_codes(naver_store, pairs)
        db.session.add(SyncLog(
            naver_store_id=store_id,
            action="PUSH_CODES",
            result="success",
            detail=f"성공 {result['success_count']} / 실패 {result['fail_count']}",
        ))
        db.session.commit()
        return jsonify(result)
    except Exception as e:
        db.session.add(SyncLog(naver_store_id=store_id, action="PUSH_CODES", result="error", detail=str(e)))
        db.session.commit()
        return jsonify({"error": str(e)}), 500


@store_bp.route("/stores/list-json")
@login_required
def stores_list_json():
    stores = _all_stores()
    return jsonify([{"id": s.id, "store_name": s.store_name} for s in stores])


@store_bp.route("/logs")
@login_required
def logs_page():
    from app.execution_logs.models import CollectionRun
    logs = SyncLog.query.order_by(SyncLog.created_at.desc()).limit(200).all()
    collection_runs = CollectionRun.query.order_by(CollectionRun.started_at.desc()).limit(100).all()
    return render_template("logs.html", logs=logs, collection_runs=collection_runs)


@store_bp.route("/logs/live")
@login_required
def logs_live_page():
    from app.execution_logs.models import CollectionRun
    running = CollectionRun.query.filter_by(status="running").order_by(CollectionRun.started_at.desc()).all()
    return render_template("log_live.html", running=running)


@store_bp.route("/logs/live-stream")
@login_required
def logs_live_stream():
    import time, json
    from app import log_buffer

    since_str = request.args.get("since", "0")
    try:
        since = float(since_str)
    except ValueError:
        since = 0.0

    def generate():
        nonlocal since
        while True:
            entries = log_buffer.get_since(since)
            for e in entries:
                since = e["t"]
                yield f"data: {json.dumps({'t': e['t'], 'msg': e['msg']}, ensure_ascii=False)}\n\n"
            time.sleep(1)

    resp = current_app.response_class(
        generate(),
        mimetype="text/event-stream",
    )
    resp.headers["Cache-Control"] = "no-cache"
    resp.headers["X-Accel-Buffering"] = "no"
    return resp


@store_bp.route("/api/running-collections")
@login_required
def api_running_collections():
    from app.execution_logs.models import CollectionRun
    running = CollectionRun.query.filter_by(status="running").order_by(CollectionRun.started_at.desc()).all()
    result = []
    for r in running:
        result.append({
            "name": r.wholesaler.name if r.wholesaler else "알 수 없음",
            "started_at": r.started_at.strftime("%H:%M:%S") if r.started_at else None,
        })
    return jsonify({"running": result})


@store_bp.route("/store-products/<int:product_id>/edit-form")
@login_required
def store_product_edit_form(product_id):
    p = StoreProduct.query.get_or_404(product_id)
    store = NaverStore.query.get_or_404(p.naver_store_id)
    if not p.origin_product_no:
        return jsonify({"error": "origin_product_no 없음"}), 400
    try:
        from store.naver.products import get_origin_product
        data = get_origin_product(p.origin_product_no, store.client_id, store.client_secret)
        origin = data.get("originProduct", {})
        return jsonify({
            "product_id": p.id,
            "origin_product_no": p.origin_product_no,
            "name": origin.get("name", p.product_name or ""),
            "sale_price": origin.get("salePrice", p.sale_price or 0),
            "status_type": origin.get("statusType", p.store_status or "SALE"),
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@store_bp.route("/store-products/<int:product_id>/edit", methods=["POST"])
@login_required
def store_product_edit(product_id):
    p = StoreProduct.query.get_or_404(product_id)
    store = NaverStore.query.get_or_404(p.naver_store_id)
    if not p.origin_product_no:
        return jsonify({"error": "origin_product_no 없음"}), 400

    new_name = request.form.get("name", "").strip()
    new_price = request.form.get("sale_price", type=int)
    new_status = request.form.get("status_type", "").strip()

    try:
        from store.naver.products import get_origin_product, update_origin_product
        data = get_origin_product(p.origin_product_no, store.client_id, store.client_secret)
        origin = data.get("originProduct", {})

        if new_name:
            origin["name"] = new_name
        if new_price:
            origin["salePrice"] = new_price
        if new_status:
            origin["statusType"] = new_status

        payload = {"originProduct": origin, "smartstoreChannelProduct": data.get("smartstoreChannelProduct", {})}
        update_origin_product(p.origin_product_no, payload, store.client_id, store.client_secret)

        # 로컬 DB 동기화
        if new_name:
            p.product_name = new_name
        if new_price:
            p.sale_price = new_price
        if new_status:
            p.store_status = new_status
        db.session.commit()

        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@store_bp.route("/store-products")
@login_required
def store_products_page():
    naver_store_id = request.args.get("naver_store_id", type=int)
    status_filter = request.args.get("status", "")
    page = request.args.get("page", 1, type=int)
    per_page = 100

    stores = _all_stores()

    query = StoreProduct.query
    if naver_store_id:
        query = query.filter_by(naver_store_id=naver_store_id)
    if status_filter:
        query = query.filter_by(store_status=status_filter)

    # 상태별 카운트
    base_query = StoreProduct.query
    if naver_store_id:
        base_query = base_query.filter_by(naver_store_id=naver_store_id)

    total = base_query.count()
    count_rows = base_query.with_entities(
        StoreProduct.store_status, func.count(StoreProduct.id)
    ).group_by(StoreProduct.store_status).all()
    counts = {code: 0 for code in STATUS_LABELS}
    for status, cnt in count_rows:
        if status in counts:
            counts[status] = cnt

    pagination = query.order_by(StoreProduct.id.desc()).paginate(page=page, per_page=per_page, error_out=False)

    return render_template(
        "store_products.html",
        stores=stores,
        selected_store_id=naver_store_id,
        status_filter=status_filter,
        status_labels=STATUS_LABELS,
        total=total,
        counts=counts,
        pagination=pagination,
        products=pagination.items,
    )


# ─────────────────────────────────────────────────────────────────────────────
# 미매칭 스토어 상품 검수 페이지 (조회 + 엑셀 다운로드)
# ─────────────────────────────────────────────────────────────────────────────

GROUP_LABELS = {
    "empty":      "코드 없음",
    "active":     "활성 도매처 (코드 정정 필요)",
    "deprecated": "폐기 도매처",
    "unknown":    "미등록 도매처",
}


def _build_unmatched_query(naver_store_id, status_filter, search_q):
    """공통 쿼리 빌더 — 페이지 / 엑셀에서 동시 사용."""
    q = StoreProduct.query.filter(StoreProduct.master_product_id.is_(None))
    if naver_store_id:
        q = q.filter(StoreProduct.naver_store_id == naver_store_id)
    if status_filter:
        q = q.filter(StoreProduct.store_status == status_filter)
    if search_q:
        like = f"%{search_q}%"
        q = q.filter(
            db.or_(
                StoreProduct.seller_management_code.ilike(like),
                StoreProduct.product_name.ilike(like),
            )
        )
    return q


def _enrich_with_group(rows, active_prefixes, prefix_to_name):
    """SQLAlchemy 결과 행에 group / suspected 정보를 부착해 반환."""
    out = []
    for sp in rows:
        group, prefix = _classify_unmatched(sp.seller_management_code or "", active_prefixes)
        suspected = prefix_to_name.get(prefix) if prefix else None
        out.append({
            "sp": sp,
            "group": group,
            "prefix": prefix,
            "suspected": suspected,
        })
    return out


@store_bp.route("/unmatched-store-products")
@login_required
def unmatched_store_products_page():
    naver_store_id = request.args.get("naver_store_id", type=int)
    status_filter = request.args.get("status", "")
    group_filter = request.args.get("group", "")
    search_q = request.args.get("q", "").strip()
    page = request.args.get("page", 1, type=int)
    per_page = 100

    stores = _all_stores()
    active_prefixes = _active_prefixes()
    prefix_to_name = _prefix_to_wholesaler_name()

    base_q = _build_unmatched_query(naver_store_id, status_filter, search_q)

    # 그룹별 카운트 — 미매칭 전체에서 그룹 분류 후 집계 (SQL 단계로 prefix 분류 어려워 Python 집계)
    # 큰 데이터셋에서도 seller_management_code + 약간의 메타만 들고와서 카운트 — 성능 OK
    count_rows = base_q.with_entities(StoreProduct.seller_management_code).all()
    group_counts = {"empty": 0, "active": 0, "deprecated": 0, "unknown": 0}
    for (code,) in count_rows:
        g, _ = _classify_unmatched(code or "", active_prefixes)
        group_counts[g] = group_counts.get(g, 0) + 1
    total_unmatched = sum(group_counts.values())

    # 그룹 필터를 적용한 페이지네이션 — 그룹은 SQL 직접 표현이 어려워서 in-memory 페이지네이션 사용
    if group_filter:
        # 그룹 필터 + 페이지 처리
        all_rows = base_q.order_by(StoreProduct.id.desc()).all()
        all_enriched = _enrich_with_group(all_rows, active_prefixes, prefix_to_name)
        filtered = [e for e in all_enriched if e["group"] == group_filter]
        total_filtered = len(filtered)
        # 단순 슬라이스 페이지네이션
        start = (page - 1) * per_page
        end = start + per_page
        items = filtered[start:end]
        # pagination 유사 객체
        class _P:
            pass
        pagination = _P()
        pagination.page = page
        pagination.per_page = per_page
        pagination.total = total_filtered
        pagination.pages = max(1, (total_filtered + per_page - 1) // per_page)
        pagination.has_prev = page > 1
        pagination.has_next = end < total_filtered
        pagination.prev_num = page - 1 if page > 1 else None
        pagination.next_num = page + 1 if end < total_filtered else None
        def _iter_pages(left_edge=2, right_edge=2, left_current=2, right_current=2):
            yield from range(1, pagination.pages + 1)
        pagination.iter_pages = _iter_pages
    else:
        sa_pagination = base_q.order_by(StoreProduct.id.desc()).paginate(
            page=page, per_page=per_page, error_out=False
        )
        items = _enrich_with_group(sa_pagination.items, active_prefixes, prefix_to_name)
        pagination = sa_pagination

    return render_template(
        "unmatched_store_products.html",
        stores=stores,
        selected_store_id=naver_store_id,
        status_filter=status_filter,
        group_filter=group_filter,
        search_q=search_q,
        status_labels=STATUS_LABELS,
        group_labels=GROUP_LABELS,
        group_counts=group_counts,
        total_unmatched=total_unmatched,
        items=items,
        pagination=pagination,
    )


@store_bp.route("/unmatched-store-products/export")
@login_required
def unmatched_store_products_export():
    """미매칭 상품을 현재 필터 기준으로 xlsx 다운로드.
    엑셀 컬럼은 사용자가 도매처 코드 정정 후 다시 업로드할 때 그대로 쓸 수 있도록
    빈 작업 컬럼 3개(새_도매처코드 / 처리방향 / 비고) 포함.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    naver_store_id = request.args.get("naver_store_id", type=int)
    status_filter = request.args.get("status", "")
    group_filter = request.args.get("group", "")
    search_q = request.args.get("q", "").strip()

    active_prefixes = _active_prefixes()
    prefix_to_name = _prefix_to_wholesaler_name()

    base_q = _build_unmatched_query(naver_store_id, status_filter, search_q)
    rows = base_q.order_by(StoreProduct.naver_store_id.asc(), StoreProduct.id.asc()).all()
    enriched = _enrich_with_group(rows, active_prefixes, prefix_to_name)
    if group_filter:
        enriched = [e for e in enriched if e["group"] == group_filter]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "미매칭 상품"

    headers = [
        "store_product_id",
        "store_name",
        "seller_management_code",
        "product_name",
        "store_status",
        "sale_price",
        "origin_product_no",
        "channel_product_no",
        "last_synced_at",
        "prefix_group",
        "추정_도매처",
        "새_도매처코드",     # 사용자 입력
        "처리방향",          # 사용자 입력 (매칭/삭제/유지 등)
        "비고",              # 사용자 입력
    ]
    ws.append(headers)
    # 헤더 스타일
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="FFE5E7EB", end_color="FFE5E7EB", fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    for e in enriched:
        sp = e["sp"]
        store_name = sp.naver_store.store_name if sp.naver_store else "-"
        last_synced = sp.last_synced_at.strftime("%Y-%m-%d %H:%M:%S") if sp.last_synced_at else ""
        ws.append([
            sp.id,
            store_name,
            sp.seller_management_code or "",
            sp.product_name or "",
            sp.store_status or "",
            sp.sale_price if sp.sale_price is not None else "",
            sp.origin_product_no if sp.origin_product_no is not None else "",
            sp.channel_product_no if sp.channel_product_no is not None else "",
            last_synced,
            GROUP_LABELS.get(e["group"], e["group"]),
            e["suspected"] or "",
            "",  # 새_도매처코드
            "",  # 처리방향
            "",  # 비고
        ])

    # 컬럼 폭 자동 조정 (간단)
    widths = [12, 12, 28, 50, 12, 10, 14, 14, 18, 22, 12, 24, 14, 24]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i) if i <= 26 else "A" + chr(64 + i - 26)].width = w

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)

    from datetime import datetime
    fname_parts = ["unmatched"]
    if naver_store_id:
        store = NaverStore.query.get(naver_store_id)
        if store:
            fname_parts.append(store.store_name.replace("/", "_"))
    if group_filter:
        fname_parts.append(group_filter)
    if status_filter:
        fname_parts.append(status_filter)
    fname_parts.append(datetime.now().strftime("%Y%m%d_%H%M%S"))
    filename = "_".join(fname_parts) + ".xlsx"

    return send_file(
        bio,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )
