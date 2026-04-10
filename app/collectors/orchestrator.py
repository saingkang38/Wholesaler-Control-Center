import logging
from app.infrastructure import db
from app.utils import kst_now
from app.execution_logs.models import CollectionRun
from app.wholesalers.models import Wholesaler

logger = logging.getLogger(__name__)


def _build_registry():
    from collectors.ownerclan import OwnerclanCollector
    from collectors.jtckorea import JtckoreaCollector
    from collectors.metaldiy import MetaldiyCollector
    from collectors.ds1008 import Ds1008Collector
    from collectors.hitdesign import HitdesignCollector
    from collectors.chingudome import ChingudomeCollector
    from collectors.zentrade import ZentraldeCollector
    from collectors.mro3 import Mro3Collector
    from collectors.feelwoo import FeelwooCollector
    from collectors.sikjaje import SikjajeCollector
    from collectors.onch3 import Onch3Collector
    from collectors.dometopia import DometopiaCollector
    return {
        "ownerclan": OwnerclanCollector,
        "jtckorea": JtckoreaCollector,
        "metaldiy": MetaldiyCollector,
        "ds1008": Ds1008Collector,
        "hitdesign": HitdesignCollector,
        "chingudome": ChingudomeCollector,
        "zentrade": ZentraldeCollector,
        "mro3": Mro3Collector,
        "feelwoo": FeelwooCollector,
        "sikjaje": SikjajeCollector,
        "onch3": Onch3Collector,
        "dometopia": DometopiaCollector,
    }


def _save_desktop_xlsx(wholesaler_code: str, wholesaler_prefix: str, items: list):
    """표준 A~AW 컬럼 레이아웃으로 저장."""
    import re
    import openpyxl
    from openpyxl.styles import PatternFill, Font
    from pathlib import Path

    if not items:
        return

    try:
        from app.settings import apply_margin

        desktop = Path(__file__).resolve().parents[2] / "downloads" / wholesaler_code
        desktop.mkdir(parents=True, exist_ok=True)

        STATUS_KO = {"active": "정상", "out_of_stock": "품절", "discontinued": "단종", "unknown": "알수없음"}

        def _fee_type(item):
            fee = item.get("shipping_fee")
            cond = item.get("shipping_condition")
            if fee == 0:
                return "FREE"
            if fee and fee > 0 and cond:
                return "CONDITIONAL_FREE"
            if fee and fee > 0:
                return "CHARGE"
            return ""

        def _cond_amount(cond_str):
            """'30,000원 이상' → 30000"""
            if not cond_str:
                return ""
            nums = re.findall(r"[\d,]+", str(cond_str))
            if nums:
                try:
                    return int(nums[0].replace(",", ""))
                except ValueError:
                    pass
            return ""

        def _v(val):
            if val is None:
                return ""
            if isinstance(val, (list, dict)):
                import json
                return json.dumps(val, ensure_ascii=False)
            return val

        HEADERS = [
            "A.도매처코드", "B.도매처상품코드", "C.판매자관리코드", "D.내부관리코드",
            "E.도매처상품명", "F.판매상품명", "G.도매처카테고리", "H.네이버카테고리ID",
            "I.매입가", "J.도매판매가", "K.마진적용가", "L.정가(설정판매가)", "M.즉시할인금액", "N.과세유형",
            "O.재고수량", "P.상품상태코드", "Q.상품상태(신상중고)",
            "R.대표이미지URL",
            "S.추가이미지1", "T.추가이미지2", "U.추가이미지3", "V.추가이미지4", "W.추가이미지5",
            "X.원산지", "Y.브랜드", "Z.제조사", "AA.모델명", "AB.키워드", "AC.인증정보", "AD.도매처상품URL",
            "AE.배송방법", "AF.배송비유형", "AG.기본배송비", "AH.무료배송조건금액",
            "AI.반품배송비", "AJ.교환배송비", "AK.출고소요일", "AL.택배사",
            "AM.출고지코드", "AN.반품교환지코드",
            "AO.옵션속성명", "AP.옵션명목록", "AQ.옵션가차액", "AR.옵션재고",
            "AS.상세설명HTML",
            "AT.수집일시", "AU.최초수집일", "AV.최종수집일", "AW.연속미수집일",
        ]

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = wholesaler_code

        # 헤더 행
        ws.append(HEADERS)
        header_fill = PatternFill("solid", fgColor="2C3E50")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.column_letter  # touch to init

        # 빈칸 열 음영 (나중에 채워야 하는 열)
        blank_cols = {4, 6, 8, 12, 13,           # D F H L M
                      19, 20, 21, 22, 23,          # S T U V W
                      25, 26, 27, 28, 29,          # Y Z AA AB AC
                      31, 37, 38, 39, 40, 41, 42,  # AE AI AJ AK AL AM AN
                      43}                           # AO
        blank_fill = PatternFill("solid", fgColor="F5F5F5")

        now_str = kst_now().strftime("%Y-%m-%d %H:%M")

        for item in items:
            extra = item.get("extra") or {}
            src_code = _v(item.get("source_product_code"))
            mgmt_code = f"{wholesaler_prefix}{src_code}" if src_code else ""
            price = item.get("price")
            margin_price = apply_margin(price) if price else ""

            row = [
                wholesaler_code,                                        # A
                src_code,                                               # B
                mgmt_code,                                              # C
                "",                                                     # D 내부관리코드 (수동)
                _v(item.get("product_name")),                           # E
                "",                                                     # F 판매상품명 (수동)
                _v(item.get("category_name")),                          # G
                "",                                                     # H 네이버카테고리ID (수동)
                _v(item.get("supply_price")),                           # I 매입가
                _v(price),                                              # J 도매판매가
                margin_price,                                           # K 마진적용가
                "",                                                     # L 정가 (옵션상품용, 수동or자동)
                "",                                                     # M 즉시할인금액
                _v(item.get("tax_type") or extra.get("과세여부", "taxable")),  # N 과세유형
                _v(item.get("stock_qty")),                              # O 재고
                STATUS_KO.get(str(item.get("status", "")), _v(item.get("status"))),  # P
                "NEW",                                                  # Q 신상/중고 기본값
                _v(item.get("image_url")),                              # R 대표이미지
                _v(extra.get("추가이미지1") or extra.get("additional_image_1")),  # S
                _v(extra.get("추가이미지2") or extra.get("additional_image_2")),  # T
                _v(extra.get("추가이미지3") or extra.get("additional_image_3")),  # U
                _v(extra.get("추가이미지4") or extra.get("additional_image_4")),  # V
                _v(extra.get("추가이미지5") or extra.get("additional_image_5")),  # W
                _v(item.get("origin")),                                 # X 원산지
                _v(item.get("brand_name") or extra.get("브랜드")),      # Y 브랜드
                _v(item.get("manufacturer") or extra.get("제조사")),    # Z 제조사
                _v(item.get("model_name") or extra.get("모델명")),      # AA 모델명
                _v(item.get("keywords") or extra.get("키워드")),        # AB 키워드
                _v(item.get("certification") or extra.get("인증정보")), # AC 인증정보
                _v(item.get("product_url") or item.get("detail_url")),  # AD 상품URL
                _v(item.get("delivery_type")),                          # AE 배송방법
                _fee_type(item),                                        # AF 배송비유형 (추론)
                _v(item.get("shipping_fee")),                           # AG 기본배송비
                _cond_amount(item.get("shipping_condition")),           # AH 무료조건금액
                _v(item.get("return_fee")),                             # AI 반품배송비
                _v(item.get("exchange_fee")),                           # AJ 교환배송비
                _v(item.get("shipping_days")),                          # AK 출고소요일
                _v(item.get("delivery_company")),                       # AL 택배사
                _v(item.get("outbound_place_code")),                    # AM 출고지코드
                _v(item.get("return_address_code")),                    # AN 반품지코드
                _v(extra.get("옵션속성명")),                            # AO 옵션속성명
                _v(extra.get("옵션")),                                  # AP 옵션명목록
                _v(extra.get("옵션가")),                                # AQ 옵션가차액
                _v(extra.get("옵션재고")),                              # AR 옵션재고
                _v(item.get("detail_description")),                     # AS 상세설명
                now_str,                                                 # AT 수집일시
                "",                                                     # AU 최초수집일 (DB)
                "",                                                     # AV 최종수집일 (DB)
                "",                                                     # AW 연속미수집일 (DB)
            ]
            ws.append(row)

            # 빈칸 열 음영 처리
            row_idx = ws.max_row
            for col_idx in blank_cols:
                ws.cell(row=row_idx, column=col_idx).fill = blank_fill

        # 열 너비 자동 조정 (헤더 기준)
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

        timestamp = kst_now().strftime("%Y%m%d_%H%M%S")
        path = desktop / f"{wholesaler_code}_{timestamp}.xlsx"
        wb.save(str(path))
        logger.info(f"[orchestrator] 엑셀 저장: {path} ({len(items)}건)")
    except Exception as e:
        logger.warning(f"[orchestrator] 엑셀 저장 실패 (무시): {e}")


def _save_raw_json(wholesaler_code: str, items: list):
    import json
    from pathlib import Path

    save_dir = Path.home() / "OneDrive" / "supplier_sync" / wholesaler_code
    save_dir.mkdir(parents=True, exist_ok=True)
    timestamp = kst_now().strftime("%Y%m%d_%H%M%S")
    path = save_dir / f"{wholesaler_code}_{timestamp}.json"
    try:
        with open(path, "w", encoding="utf-8") as f:
            json.dump(items, f, ensure_ascii=False, indent=2)
        logger.info(f"[orchestrator] 원본 저장: {path} ({len(items)}건)")
    except Exception as e:
        logger.warning(f"[orchestrator] 원본 저장 실패 (무시): {e}")


def run_collection(wholesaler_code: str, trigger_type: str = "manual", user_id: int = None):
    wholesaler = Wholesaler.query.filter_by(code=wholesaler_code, is_active=True).first()
    if not wholesaler:
        return {"success": False, "error": f"도매처를 찾을 수 없음: {wholesaler_code}"}

    registry = _build_registry()
    collector_class = registry.get(wholesaler_code)
    if not collector_class:
        return {"success": False, "error": f"{wholesaler_code} collector 미등록"}

    # 설정 미완료 여부를 미리 확인 (CollectionRun 생성 전)
    _CONFIG_KEYWORDS = ("미설정", "환경변수 없음", "환경변수없음", "not configured", "LOGIN_ID", "LOGIN_PASSWORD")

    run = CollectionRun(
        wholesaler_id=wholesaler.id,
        trigger_type=trigger_type,
        status="running",
        started_at=kst_now(),
        created_by_user_id=user_id,
    )
    db.session.add(run)
    db.session.commit()

    master_stats = {}  # try 블록 외부 선언 — 예외 발생 시에도 반환값 안전
    try:
        collector = collector_class()
        result = collector.run()

        # 설정 미완료로 수집 자체를 건너뛴 경우 — 알림 없이 로그만
        error_msg = result.get("error_summary") or result.get("error") or ""
        if not result.get("success") and any(kw in error_msg for kw in _CONFIG_KEYWORDS):
            run.status = "skipped"
            run.error_summary = error_msg
            run.finished_at = kst_now()
            db.session.commit()
            logger.info(f"[orchestrator] {wholesaler_code} 설정 미완료 — 수집 건너뜀: {error_msg}")
            return {"success": False, "not_configured": True, "error": error_msg}

        run.status = "success" if result.get("success") else "failed"
        run.total_items = result.get("total_items", 0)
        run.total_pages = result.get("total_pages", 0)
        run.success_count = result.get("success_count", 0)
        run.fail_count = result.get("fail_count", 0)
        run.error_summary = result.get("error_summary")

        if result.get("success") and result.get("items"):
            _save_raw_json(wholesaler_code, result["items"])
            _save_desktop_xlsx(wholesaler_code, wholesaler.prefix or "", result["items"])
            from app.normalization import save_normalized_products
            from app.master import process_master_update
            saved = save_normalized_products(
                wholesaler_id=wholesaler.id,
                run_id=run.id,
                items=result["items"]
            )
            logger.info(f"[orchestrator] 저장 완료: {saved}건")
            master_stats = process_master_update(wholesaler.id, result["items"])
            logger.info(f"[orchestrator] 마스터 업데이트 완료")
            from app.actions import detect_action_signals
            detect_action_signals(wholesaler.id)
            logger.info(f"[orchestrator] 액션 시그널 갱신 완료")

    except Exception as e:
        logger.error(f"[orchestrator] 오류 발생: {e}")
        run.status = "failed"
        run.error_summary = str(e)
        result = {"success": False, "error": str(e)}

    finally:
        run.finished_at = kst_now()
        db.session.commit()

    return {
        "success": run.status == "success",
        "run_id": run.id,
        "total_items": run.total_items,
        "master_stats": master_stats,
        "error": run.error_summary,
    }
