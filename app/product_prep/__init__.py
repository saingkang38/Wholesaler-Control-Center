import io
import logging
import threading
from flask import Blueprint, render_template, request, jsonify, Response
from flask_login import login_required
from app.master.models import MasterProduct
from app.infrastructure import db

product_prep_bp = Blueprint("product_prep", __name__)
logger = logging.getLogger(__name__)


@product_prep_bp.route("/product-prep")
@login_required
def product_prep_page():
    from app.wholesalers.models import Wholesaler
    wholesalers = Wholesaler.query.order_by(Wholesaler.name).all()
    wholesaler_id = request.args.get("wholesaler_id", type=int)
    limit = request.args.get("limit", 1000, type=int)
    if limit not in (0, 20, 50, 100, 300, 500, 1000, 2000):
        limit = 1000
    page = request.args.get("page", 1, type=int)
    if page < 1:
        page = 1

    q = MasterProduct.query.filter(
        MasterProduct.is_prep_ready == False,
        MasterProduct.current_status == "active",
    )
    if wholesaler_id:
        q = q.filter(MasterProduct.wholesaler_id == wholesaler_id)

    total = q.count()
    import math
    if limit == 0:
        total_pages = 1
        page = 1
        products = q.order_by(MasterProduct.id.desc()).all()
    else:
        total_pages = max(1, math.ceil(total / limit))
        if page > total_pages:
            page = total_pages
        products = q.order_by(MasterProduct.id.desc()).offset((page - 1) * limit).limit(limit).all()

    return render_template(
        "product_prep.html",
        products=products,
        wholesalers=wholesalers,
        selected_wholesaler_id=wholesaler_id,
        selected_limit=limit,
        page=page,
        total_pages=total_pages,
        total=total,
    )


def _download_images_bg(image_pairs: list, image_dir: str):
    """백그라운드에서 이미지 다운로드 — (supplier_code, image_url) 리스트"""
    import requests
    from pathlib import Path

    save_dir = Path(image_dir)
    save_dir.mkdir(parents=True, exist_ok=True)
    ok = 0
    fail = 0
    for code, url in image_pairs:
        if not url:
            continue
        try:
            ext = url.split("?")[0].rsplit(".", 1)[-1][:5].lower()
            if ext not in ("jpg", "jpeg", "png", "gif", "webp"):
                ext = "jpg"
            filename = f"{code}.{ext}"
            path = save_dir / filename
            if path.exists():
                ok += 1
                continue
            r = requests.get(url, timeout=10)
            if r.status_code == 200:
                path.write_bytes(r.content)
                ok += 1
            else:
                fail += 1
        except Exception:
            fail += 1
    logger.info(f"[product_prep] 이미지 다운로드 완료 — 성공 {ok} / 실패 {fail}")


@product_prep_bp.route("/product-prep/download", methods=["GET", "POST"])
@login_required
def download_xlsx():
    import subprocess
    import openpyxl
    from pathlib import Path
    from datetime import datetime
    from openpyxl.styles import Font, PatternFill, Alignment
    from app.settings.models import PrepSetting

    # POST: JSON body {ids: [...], wholesaler_id: ...}
    # GET:  ?wholesaler_id=...  (기존 호환)
    body = request.get_json(silent=True) or {}
    ids = body.get("ids")  # list of string IDs, or None
    wholesaler_id = body.get("wholesaler_id") or request.args.get("wholesaler_id", type=int)

    if ids:
        # 선택된 항목만
        int_ids = [int(i) for i in ids if str(i).isdigit()]
        products = MasterProduct.query.filter(MasterProduct.id.in_(int_ids)).order_by(MasterProduct.id.desc()).all()
    else:
        q = MasterProduct.query.filter(
            MasterProduct.is_prep_ready == False,
            MasterProduct.current_status == "active",
        )
        if wholesaler_id:
            q = q.filter(MasterProduct.wholesaler_id == wholesaler_id)
        products = q.order_by(MasterProduct.id.desc()).all()

    setting = PrepSetting.get()
    base_excel_dir = Path(setting.excel_dir) if setting.excel_dir else Path.home() / "Desktop" / "상품가공"

    # 도매처명 폴더 생성
    if wholesaler_id:
        from app.wholesalers.models import Wholesaler
        w = Wholesaler.query.get(wholesaler_id)
        folder_name = w.name if w else str(wholesaler_id)
    elif ids:
        # 선택 항목의 도매처가 섞여 있을 수 있으므로 "선택상품" 폴더 사용
        folder_name = "선택상품"
    else:
        folder_name = "전체"

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    excel_dir = base_excel_dir / f"{folder_name}_{timestamp}"
    image_dir = Path(setting.image_dir) / f"{folder_name}_{timestamp}" if setting.image_dir else excel_dir / "이미지"

    # ── 엑셀 생성 ────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "상품가공"

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    h1 = ws.cell(row=1, column=1, value="원본 상품명*")
    h1.font = Font(bold=True, color="FFFFFF")
    h1.fill = header_fill
    h1.alignment = Alignment(horizontal="center")

    h2 = ws.cell(row=1, column=2, value="상세 설명*\nhtml 코드 그대로 삽입")
    h2.font = Font(bold=True, color="FFFFFF")
    h2.fill = header_fill
    h2.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.row_dimensions[1].height = 30

    for row, m in enumerate(products, 2):
        ws.cell(row=row, column=1, value=m.product_name or "")
        ws.cell(row=row, column=2, value=m.detail_description or "")

    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 80

    excel_dir.mkdir(parents=True, exist_ok=True)
    filename = f"product_prep_{timestamp}.xlsx"
    save_path = excel_dir / filename
    wb.save(str(save_path))

    # ── 이미지 백그라운드 다운로드 ───────────────────────
    image_pairs = [(m.supplier_product_code, m.image_url) for m in products if m.image_url]
    t = threading.Thread(
        target=_download_images_bg,
        args=(image_pairs, str(image_dir)),
        daemon=True,
    )
    t.start()

    # 탐색기로 엑셀 저장 폴더 열기
    try:
        subprocess.Popen(["explorer", str(excel_dir)])
    except Exception as e:
        logger.warning(f"[product_prep] 탐색기 열기 실패 (무시): {e}")

    return jsonify({
        "success": True,
        "path": str(save_path),
        "filename": filename,
        "image_count": len(image_pairs),
        "image_dir": str(image_dir),
    })


@product_prep_bp.route("/product-prep/upload", methods=["POST"])
@login_required
def upload_xlsx():
    import openpyxl

    f = request.files.get("file")
    if not f:
        return jsonify({"error": "파일 없음"}), 400

    try:
        wb = openpyxl.load_workbook(io.BytesIO(f.read()), data_only=True)
        ws = wb.active
    except Exception as e:
        return jsonify({"error": f"xlsx 파일 읽기 실패: {e}"}), 400

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return jsonify({"error": "빈 파일"}), 400

    # 컬럼 순서: 상품명 / 이미지url / 가공된상품명 / 네이버카테고리번호
    header = [str(h).strip() if h else "" for h in rows[0]]
    try:
        col_name     = header.index("상품명")
        col_edited   = header.index("가공된상품명")
        col_category = header.index("네이버카테고리번호")
    except ValueError as e:
        return jsonify({"error": f"필수 컬럼 없음: {e} — 헤더: {header}"}), 400

    updated = 0
    skipped = 0
    errors = []
    incomplete = []

    for row in rows[1:]:
        def cell(idx):
            v = row[idx] if idx < len(row) else None
            return str(v).strip() if v is not None else ""

        product_name = cell(col_name)
        edited_name  = cell(col_edited)
        category_id  = cell(col_category)

        if not product_name:
            skipped += 1
            continue

        master = MasterProduct.query.filter_by(product_name=product_name).first()
        if not master:
            errors.append(f"DB에 없음: {product_name[:40]}")
            continue

        if edited_name and category_id:
            # 조건 충족 → 저장
            master.edited_name = edited_name
            master.category_id = category_id
            master.is_prep_ready = True
            updated += 1
        elif edited_name and not category_id:
            # 카테고리 없음 → 저장 안 함
            incomplete.append(f"카테고리 없음: {product_name[:40]}")
        elif category_id and not edited_name:
            # 상품명 없음 → 저장 안 함
            incomplete.append(f"가공된상품명 없음: {product_name[:40]}")
        else:
            skipped += 1

    db.session.commit()
    return jsonify({
        "updated": updated,
        "skipped": skipped,
        "incomplete": incomplete[:30],
        "errors": errors[:20],
    })


# ── 이미지 가공 ───────────────────────────────────────────────────────────────

_img_job = {"status": "idle", "total": 0, "done": 0, "errors": 0, "message": ""}


def _process_images_bg(image_dir: str, output_dir: str,
                        inner_scale: int, rotation: int,
                        output_size: int, quality: int):
    """백그라운드 이미지 가공 작업"""
    global _img_job
    from pathlib import Path
    try:
        from PIL import Image
    except ImportError:
        _img_job.update({"status": "error", "message": "Pillow 미설치 (pip install Pillow)"})
        return

    src_dir = Path(image_dir)
    dst_dir = Path(output_dir)
    dst_dir.mkdir(parents=True, exist_ok=True)

    exts = {".jpg", ".jpeg", ".png", ".gif", ".webp"}
    files = [f for f in src_dir.iterdir() if f.suffix.lower() in exts]
    _img_job.update({"status": "running", "total": len(files), "done": 0, "errors": 0, "message": ""})

    for f in files:
        try:
            img = Image.open(f).convert("RGB")

            # 1) 회전
            if rotation:
                img = img.rotate(-rotation, expand=True, fillcolor=(255, 255, 255))

            # 2) 출력 캔버스 크기 결정
            canvas_px = output_size if output_size else max(img.width, img.height)

            # 3) 내부 크기 적용 (inner_scale%)
            inner_px = int(canvas_px * inner_scale / 100)
            ratio = inner_px / max(img.width, img.height)
            new_w = int(img.width * ratio)
            new_h = int(img.height * ratio)
            img = img.resize((new_w, new_h), Image.LANCZOS)

            # 4) 흰 캔버스에 중앙 배치
            canvas = Image.new("RGB", (canvas_px, canvas_px), (255, 255, 255))
            offset_x = (canvas_px - new_w) // 2
            offset_y = (canvas_px - new_h) // 2
            canvas.paste(img, (offset_x, offset_y))

            # 5) 저장 (quality 적용)
            out_path = dst_dir / (f.stem + ".jpg")
            canvas.save(str(out_path), "JPEG", quality=quality, optimize=True)

            _img_job["done"] += 1
        except Exception as e:
            logger.warning(f"[process_images] {f.name} 실패: {e}")
            _img_job["errors"] += 1

    _img_job["status"] = "done"
    _img_job["message"] = f"완료 — 성공 {_img_job['done']}개 / 오류 {_img_job['errors']}개"
    logger.info(f"[product_prep] 이미지 가공 완료: {_img_job['message']}")


@product_prep_bp.route("/product-prep/process-images", methods=["POST"])
@login_required
def start_process_images():
    global _img_job
    if _img_job.get("status") == "running":
        return jsonify({"error": "이미 가공 중입니다."}), 409

    from app.settings.models import PrepSetting
    s = PrepSetting.get()

    image_dir   = s.image_dir
    output_dir  = s.processed_image_dir
    inner_scale = s.img_inner_scale or 100
    rotation    = s.img_rotation or 0
    output_size = s.img_output_size
    quality     = s.img_quality or 100

    if not image_dir or not output_dir:
        return jsonify({"error": "이미지 저장 경로 또는 가공 이미지 저장 경로가 설정되지 않았습니다."}), 400

    from pathlib import Path
    if not Path(image_dir).exists():
        return jsonify({"error": f"이미지 폴더가 없습니다: {image_dir}"}), 400

    t = threading.Thread(
        target=_process_images_bg,
        args=(image_dir, output_dir, inner_scale, rotation, output_size, quality),
        daemon=True,
    )
    t.start()
    return jsonify({"ok": True, "message": "이미지 가공 시작됨"})


@product_prep_bp.route("/product-prep/process-images/status")
@login_required
def process_images_status():
    return jsonify(_img_job)
