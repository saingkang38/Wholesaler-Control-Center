import json
import logging
import io
from flask import Blueprint, render_template, request, jsonify, send_file
from flask_login import login_required
from app.utils import kst_now
from sqlalchemy.orm import joinedload
from app.infrastructure import db
from app.actions.models import ActionSignal
from app.master.models import MasterProduct
from app.store.models import StoreProduct

actions_bp = Blueprint("actions", __name__)
logger = logging.getLogger(__name__)


SIGNAL_LABELS = {
    "PRICE_UP_NEEDED":     {"label": "가격 인상 필요", "badge": "danger"},
    "PRICE_DOWN_POSSIBLE": {"label": "가격 인하 가능", "badge": "info"},
    "SUSPEND_NEEDED":      {"label": "판매 중지 필요", "badge": "warning"},
    "RESUME_POSSIBLE":     {"label": "판매 재개 가능", "badge": "success"},
    "DISCONTINUE_NEEDED":  {"label": "단종 처리 필요", "badge": "dark"},
    "OPTION_PRICE_CHANGE": {"label": "옵션가 변동", "badge": "warning"},
    "OPTION_STOCK_CHANGE": {"label": "옵션 재고 변동", "badge": "secondary"},
    "OPTION_ADD":          {"label": "옵션 추가/변경", "badge": "primary"},
    "OPTION_STOCK_REFILL_NEEDED": {"label": "옵션 재고 복구", "badge": "info"},
    "DETAIL_CHANGE":       {"label": "상세페이지 갱신", "badge": "info"},
}


def _refresh_required_fields_meta(stored_json: str | None) -> list:
    """DB에 저장된 required_fields_missing JSON을 KNOWN_REQUIRED_FIELDS의 최신 메타와 머지.

    저장 시점에 박힌 kind/label/hint/options가 코드 변경 후에도 화면에 그대로 보이지 않도록,
    이름(name) + optional 플래그만 유지하고 메타는 최신 KNOWN 정의로 덮는다.

    - KNOWN에서 제거된 옛날 추측 키는 폴백 없이 드롭 (라우트 가드와 일관성 유지)
    - co_required는 자동 추가 — 옛 시그널도 즉시 진짜 키 모달로 갱신됨
    """
    if not stored_json:
        return []
    try:
        stored = json.loads(stored_json) or []
    except Exception:
        return []
    from app.actions.required_fields import _make_field_entry, KNOWN_REQUIRED_FIELDS
    refreshed: list = []
    seen: set = set()
    for f in stored:
        if not isinstance(f, dict):
            continue
        name = f.get("name") or ""
        if not name or name in seen:
            continue
        entry = _make_field_entry(name, f.get("message") or "")
        if entry is None:
            continue  # KNOWN 미등록 — 드롭 (옛날 추측 키 제거)
        if f.get("optional"):
            entry["optional"] = True
        seen.add(name)
        refreshed.append(entry)
        # co_required 자동 노출 — 화면에서 항상 진짜 키 그룹으로 정렬됨
        meta = KNOWN_REQUIRED_FIELDS.get(name) or {}
        for co_name in meta.get("co_required") or []:
            if co_name in seen:
                continue
            co_entry = _make_field_entry(co_name)
            if not co_entry:
                continue
            co_entry["optional"] = True
            seen.add(co_name)
            refreshed.append(co_entry)
    return refreshed


@actions_bp.route("/actions")
@login_required
def actions_page():
    status_filter = request.args.get("status", "pending")
    per_page = request.args.get("per_page", 50, type=int)
    page = request.args.get("page", 1, type=int)

    valid_per_page = [30, 50, 100, 300, 500, 1000, 0]
    if per_page not in valid_per_page:
        per_page = 50

    store_filter = request.args.get("store_id", 0, type=int)
    signal_type_filter = request.args.get("signal_type", "")
    option_type_filter = request.args.get("option_type", "no_option")
    option_add_kind_filter = request.args.get("option_add_kind", "")  # "new" | "existing" | ""
    failure_kind_filter = request.args.get("failure_kind", "")  # "tag_blocked" | ""
    search_q = request.args.get("q", "").strip()

    from app.store.models import StoreProduct, NaverStore
    query = ActionSignal.query.filter_by(status=status_filter)
    # failed 탭 안에서 분류별 빠른 필터
    if status_filter == "failed" and failure_kind_filter == "tag_blocked":
        # sellerTags 등록불가 단어 에러
        query = query.filter(ActionSignal.error_message.like("%등록불가인 단어(%"))
    elif status_filter == "failed" and failure_kind_filter == "consumption_date":
        # 식품 소비기한 누락 에러 (영어 필드명으로 안전하게 매칭)
        query = query.filter(ActionSignal.error_message.like("%consumptionDate%"))
    elif status_filter == "failed" and failure_kind_filter == "close_status":
        # 자동 복구 불가 상품(CLOSE/PROHIBITION/UNADMISSION) 분류.
        # 메시지 텍스트가 다양해도(예: "한도 수를 초과" 같은 misleading 메시지 포함)
        # 상품 상태 기반으로 묶어 운영자가 어드민에서 일괄 복구할 수 있게 함.
        _close_sub = (db.session.query(StoreProduct.id)
                      .filter(StoreProduct.store_status.in_(("CLOSE", "PROHIBITION", "UNADMISSION"))))
        query = query.filter(db.or_(
            ActionSignal.error_message.like("%판매상태 변경이 가능합니다%"),
            ActionSignal.store_product_id.in_(_close_sub),
        ))
    if store_filter:
        sub = db.session.query(StoreProduct.id).filter_by(naver_store_id=store_filter).subquery()
        query = query.filter(ActionSignal.store_product_id.in_(sub))
    if signal_type_filter == "PRICE":
        query = query.filter(ActionSignal.signal_type.in_(["PRICE_UP_NEEDED", "PRICE_DOWN_POSSIBLE"]))
    elif signal_type_filter == "OPTION":
        query = query.filter(ActionSignal.signal_type.in_(["OPTION_PRICE_CHANGE", "OPTION_ADD", "OPTION_STOCK_CHANGE", "OPTION_STOCK_REFILL_NEEDED"]))
    elif signal_type_filter:
        query = query.filter(ActionSignal.signal_type == signal_type_filter)
    # 실패 분류 빠른 필터 — 옵션 유형과 무관하므로 option_type 필터(기본값 'no_option') 강제 무시.
    # 안 그러면 옵션 있는 상품의 분류 결과가 모두 화면에서 사라짐.
    if status_filter == "failed" and failure_kind_filter in ("tag_blocked", "consumption_date", "close_status"):
        option_type_filter = ""
    if option_type_filter:
        query = query.join(MasterProduct, ActionSignal.master_product_id == MasterProduct.id)
        if option_type_filter == "no_option":
            query = query.filter(
                db.or_(MasterProduct.options_text == None, MasterProduct.options_text == "")
            )
        elif option_type_filter == "option_no_extra":
            query = query.filter(
                MasterProduct.options_text != None,
                MasterProduct.options_text != "",
                db.or_(
                    MasterProduct.option_diffs == None,
                    MasterProduct.option_diffs == "",
                    db.func.replace(db.func.replace(MasterProduct.option_diffs, "0", ""), "\n", "") == "",
                ),
            )
        elif option_type_filter == "option_with_extra":
            query = query.filter(
                MasterProduct.options_text != None,
                MasterProduct.options_text != "",
                MasterProduct.option_diffs != None,
                MasterProduct.option_diffs != "",
                db.func.replace(db.func.replace(MasterProduct.option_diffs, "0", ""), "\n", "") != "",
            )
    # OPTION_ADD 신규/기존 필터 (signal_type=OPTION_ADD 일 때만 의미 있음)
    if option_add_kind_filter and signal_type_filter == "OPTION_ADD":
        if option_add_kind_filter == "new":
            sub_kind = db.session.query(StoreProduct.id).filter(StoreProduct.applied_options_text.is_(None)).subquery()
        elif option_add_kind_filter == "existing":
            sub_kind = db.session.query(StoreProduct.id).filter(StoreProduct.applied_options_text.isnot(None)).subquery()
        else:
            sub_kind = None
        if sub_kind is not None:
            query = query.filter(ActionSignal.store_product_id.in_(sub_kind))
    if search_q:
        sp_sub = db.session.query(StoreProduct.id).filter(
            db.or_(
                StoreProduct.product_name.ilike(f"%{search_q}%"),
                StoreProduct.seller_management_code.ilike(f"%{search_q}%"),
            )
        ).subquery()
        query = query.filter(ActionSignal.store_product_id.in_(sp_sub))
    query = query.order_by(ActionSignal.detected_at.desc())

    all_stores = NaverStore.query.order_by(NaverStore.store_name).all()

    if per_page == 0:
        signals = query.all()
        pagination = None
        total = len(signals)
    else:
        pagination = query.paginate(page=page, per_page=per_page, error_out=False)
        signals = pagination.items
        total = pagination.total

    from app.settings import apply_margin

    def _apply_margin_cached(price):
        return apply_margin(price) if price else None

    rows = []
    for s in signals:
        current = json.loads(s.current_value) if s.current_value else {}
        suggested = json.loads(s.suggested_value) if s.suggested_value else {}

        if s.signal_type == "OPTION_PRICE_CHANGE":
            # 도매가격=도매기준가 / 마진적용=실판매가 / 판매가격=정가(설정판매가)
            wholesale_price = suggested.get("base_price")
            margin_price    = suggested.get("sale_price")      # apply_margin(base_price) 이미 계산됨
            sale_price      = suggested.get("list_price")      # 조건 충족 정가
            discount        = suggested.get("discount", 0)
            option_count    = len(suggested.get("additions", []))
        elif s.signal_type == "OPTION_ADD":
            wholesale_price = suggested.get("base_price")
            margin_price    = _apply_margin_cached(wholesale_price) if wholesale_price else None
            sale_price      = s.store.sale_price if s.store else None
            discount        = 0
            option_count    = None
        elif s.signal_type == "OPTION_STOCK_CHANGE":
            wholesale_price = None
            margin_price    = None
            sale_price      = s.store.sale_price if s.store else None
            discount        = 0
            option_count    = None
        elif s.signal_type == "OPTION_STOCK_REFILL_NEEDED":
            wholesale_price = None
            margin_price    = None
            sale_price      = s.store.sale_price if s.store else None
            discount        = 0
            option_count    = None
        elif s.signal_type in ("SUSPEND_NEEDED", "RESUME_POSSIBLE", "DISCONTINUE_NEEDED"):
            _m2             = s.master
            wholesale_price = _m2.price if _m2 else None
            margin_price    = _apply_margin_cached(wholesale_price) if wholesale_price else None
            sale_price      = s.store.sale_price if s.store else None
            discount        = 0
            option_count    = None
        else:
            s_price         = suggested.get("sale_price")
            wholesale_price = s_price if s_price is not None else current.get("sale_price")
            sale_price      = current.get("sale_price")
            margin_price    = _apply_margin_cached(wholesale_price) if wholesale_price else None
            discount        = 0
            option_count    = None

        # OPTION_ADD 시그널: 현재 스토어 옵션 vs 적용될 도매처 옵션 표시
        current_option_rows = []
        new_option_rows = []
        if s.signal_type == "OPTION_ADD":
            from app.settings import calculate_option_pricing as _calc_pricing
            # 현재 스토어 옵션 — applied 이력으로 추가금 역산, 없으면 이름만
            cur_opts_text = current.get("options_text") or ""
            if cur_opts_text:
                cur_names = [n.strip() for n in cur_opts_text.split("\n") if n.strip()]
                if s.store and s.store.applied_option_diffs and s.store.applied_option_base_price:
                    try:
                        cur_pricing = _calc_pricing(s.store.applied_option_base_price, s.store.applied_option_diffs)
                        cur_adds = cur_pricing["additions"]
                        current_option_rows = [
                            (cur_names[i], cur_adds[i] if i < len(cur_adds) else 0)
                            for i in range(len(cur_names))
                        ]
                    except Exception:
                        current_option_rows = [(n, None) for n in cur_names]
                else:
                    current_option_rows = [(n, None) for n in cur_names]
            # 적용될 새 옵션 — master에서 최신 데이터 우선, 없으면 suggested_value 폴백
            _m = s.master
            new_opts_text = (_m.options_text if _m else None) or suggested.get("options_text") or ""
            new_diffs_text = (_m.option_diffs if _m else None) or suggested.get("option_diffs") or ""
            new_base = (_m.price if _m else None) or suggested.get("base_price")
            if new_opts_text:
                names = [n.strip() for n in new_opts_text.split("\n") if n.strip()]
                raw_diffs = []
                for d in new_diffs_text.split("\n"):
                    d = d.strip()
                    try:
                        raw_diffs.append(int(d))
                    except ValueError:
                        raw_diffs.append(0)
                # 마진 적용 추가금 계산
                margin_adds = []
                if new_base:
                    try:
                        new_pricing = _calc_pricing(new_base, new_diffs_text)
                        margin_adds = new_pricing["additions"]
                    except Exception:
                        pass
                for i, name in enumerate(names):
                    raw_d = raw_diffs[i] if i < len(raw_diffs) else 0
                    margin_d = margin_adds[i] if i < len(margin_adds) else None
                    new_option_rows.append((name, raw_d, margin_d))

        # PRICE 시그널 + 옵션 상품인 경우: 승인 시 적용될 옵션별 가격 계산
        option_details = []
        store_option_rows = []  # 현재 스마트스토어 옵션별 가격
        pricing_list_price = pricing_discount = pricing_sale_price = None
        if s.signal_type in ("PRICE_UP_NEEDED", "PRICE_DOWN_POSSIBLE") and wholesale_price:
            master = s.master
            if master and master.option_diffs and master.options_text:
                try:
                    from app.settings import calculate_option_pricing
                    opt_p = calculate_option_pricing(wholesale_price, master.option_diffs)
                    names = [n.strip() for n in master.options_text.split("\n") if n.strip()]
                    adds  = opt_p["additions"]
                    base  = opt_p["sale_price"]
                    pricing_list_price = opt_p["list_price"]
                    pricing_discount   = opt_p["discount"]
                    pricing_sale_price = opt_p["sale_price"]
                    option_details = [
                        (names[i] if i < len(names) else f"옵션{i+1}", base + (adds[i] if i < len(adds) else 0))
                        for i in range(len(names))
                    ]
                    # 현재 스마트스토어 옵션 가격 계산
                    # 우선순위 1: Naver 실제 추가금 캐시 (sale_price + naver_addition)
                    if s.store and s.store.naver_cached_additions and s.store.sale_price:
                        try:
                            cur_adds_raw = [int(v) for v in s.store.naver_cached_additions.split("\n") if v.strip()]
                            store_option_rows = [
                                (names[i] if i < len(names) else f"옵션{i+1}",
                                 s.store.sale_price + (cur_adds_raw[i] if i < len(cur_adds_raw) else 0))
                                for i in range(len(names))
                            ]
                        except Exception:
                            pass
                    # 우선순위 2: 적용 이력으로 역산
                    if not store_option_rows and s.store and s.store.applied_option_base_price and s.store.applied_option_diffs:
                        try:
                            cur_p = calculate_option_pricing(s.store.applied_option_base_price, s.store.applied_option_diffs)
                            cur_base = cur_p["sale_price"]
                            cur_adds = cur_p["additions"]
                            store_option_rows = [
                                (names[i] if i < len(names) else f"옵션{i+1}", cur_base + (cur_adds[i] if i < len(cur_adds) else 0))
                                for i in range(len(names))
                            ]
                        except Exception:
                            pass
                except Exception:
                    pass
            # 옵션 상품: 마진적용가격을 실제 최저 옵션가로 교체 (정가 6,080 → 최저 옵션가 4,280)
            if pricing_sale_price:
                margin_price = pricing_sale_price

        # 판매상태 / 전시상태 파생
        raw_status = s.store.store_status if s.store else ""
        SALE_LABELS = {
            "SALE":        ("판매중",   "success"),
            "SUSPENSION":  ("판매중지", "warning"),
            "CLOSE":       ("판매종료", "dark"),
            "WAIT":        ("판매대기", "secondary"),
            "SOLDOUT":     ("품절",     "danger"),
            "PROHIBITION": ("판매금지", "danger"),
        }
        sale_label, sale_badge = SALE_LABELS.get(raw_status, ("-", "secondary"))
        if raw_status in ("SALE", "SUSPENSION", "SOLDOUT"):
            display_label, display_badge = "전시중", "info"
        elif raw_status in ("CLOSE", "WAIT", "PROHIBITION"):
            display_label, display_badge = "전시안함", "secondary"
        else:
            display_label, display_badge = "-", "secondary"

        rows.append({
            "id": s.id,
            "store_product_id": s.store_product_id,
            "signal_type": s.signal_type,
            "label": SIGNAL_LABELS.get(s.signal_type, {}).get("label", s.signal_type),
            "badge": SIGNAL_LABELS.get(s.signal_type, {}).get("badge", "secondary"),
            "wholesaler_name": s.master.wholesaler.name if s.master and s.master.wholesaler else "-",
            "store_name": s.store.naver_store.store_name if s.store and s.store.naver_store else "-",
            "product_name": s.master.product_name if s.master else "-",
            "seller_code": s.store.seller_management_code if s.store else "-",
            "channel_product_no": s.store.channel_product_no if s.store else None,
            "wholesale_price": wholesale_price,
            "margin_price": margin_price,
            "sale_price": sale_price,
            "discount": discount,
            "option_count": option_count,
            "option_details": option_details,
            "store_option_rows": store_option_rows,
            "pricing_list_price": pricing_list_price,
            "pricing_discount":   pricing_discount,
            "pricing_sale_price": pricing_sale_price,
            "current_option_rows": current_option_rows,
            "new_option_rows": new_option_rows,
            "detected_at": s.detected_at.strftime("%Y-%m-%d %H:%M") if s.detected_at else "-",
            "status": s.status,
            "error_message": s.error_message,
            "required_fields_missing": s.required_fields_missing,
            # 시그널이 awaiting_input으로 들어간 시점의 메타가 DB에 박혀 있지만,
            # 코드(KNOWN_REQUIRED_FIELDS) 변경이 즉시 화면에 반영되도록 매번 fresh 머지.
            # 키 이름과 optional 플래그만 유지하고 kind/label/hint/options는 최신 메타로 덮음.
            "required_fields_list": _refresh_required_fields_meta(s.required_fields_missing),
            "sale_label": sale_label,
            "sale_badge": sale_badge,
            "display_label": display_label,
            "display_badge": display_badge,
            "is_new_option": (s.signal_type == "OPTION_ADD" and s.store and not s.store.applied_options_text),
        })
    pending_count = ActionSignal.query.filter_by(status="pending").count()
    failed_count = ActionSignal.query.filter_by(status="failed").count()
    awaiting_input_count = ActionSignal.query.filter_by(status="awaiting_input").count()
    tag_blocked_count = (
        ActionSignal.query
        .filter_by(status="failed")
        .filter(ActionSignal.error_message.like("%등록불가인 단어(%"))
        .count()
    )
    consumption_date_count = (
        ActionSignal.query
        .filter_by(status="failed")
        .filter(ActionSignal.error_message.like("%consumptionDate%"))
        .count()
    )
    _close_sub_for_count = (
        db.session.query(StoreProduct.id)
        .filter(StoreProduct.store_status.in_(("CLOSE", "PROHIBITION", "UNADMISSION")))
    )
    close_status_count = (
        ActionSignal.query
        .filter_by(status="failed")
        .filter(db.or_(
            ActionSignal.error_message.like("%판매상태 변경이 가능합니다%"),
            ActionSignal.store_product_id.in_(_close_sub_for_count),
        ))
        .count()
    )

    def _option_type_count(otype):
        base = ActionSignal.query.filter_by(status="pending").join(
            MasterProduct, ActionSignal.master_product_id == MasterProduct.id
        )
        if otype == "no_option":
            return base.filter(
                db.or_(MasterProduct.options_text == None, MasterProduct.options_text == "")
            ).count()
        elif otype == "option_no_extra":
            return base.filter(
                MasterProduct.options_text != None,
                MasterProduct.options_text != "",
                db.or_(
                    MasterProduct.option_diffs == None,
                    MasterProduct.option_diffs == "",
                    db.func.replace(db.func.replace(MasterProduct.option_diffs, "0", ""), "\n", "") == "",
                ),
            ).count()
        elif otype == "option_with_extra":
            return base.filter(
                MasterProduct.options_text != None,
                MasterProduct.options_text != "",
                MasterProduct.option_diffs != None,
                MasterProduct.option_diffs != "",
                db.func.replace(db.func.replace(MasterProduct.option_diffs, "0", ""), "\n", "") != "",
            ).count()

    no_option_count = _option_type_count("no_option")
    option_no_extra_count = _option_type_count("option_no_extra")
    option_with_extra_count = _option_type_count("option_with_extra")

    # OPTION_ADD 신규/기존 카운트 (검수 흐름 안내용)
    opt_add_new_count = (
        ActionSignal.query.filter_by(status="pending", signal_type="OPTION_ADD")
        .join(StoreProduct, ActionSignal.store_product_id == StoreProduct.id)
        .filter(StoreProduct.applied_options_text.is_(None)).count()
    )
    opt_add_existing_count = (
        ActionSignal.query.filter_by(status="pending", signal_type="OPTION_ADD")
        .join(StoreProduct, ActionSignal.store_product_id == StoreProduct.id)
        .filter(StoreProduct.applied_options_text.isnot(None)).count()
    )

    # 도매처별 OPTION_ADD 카운트 (option_add_kind 필터 활성 시 일괄 승인 패널용)
    wholesaler_bulk_counts = []
    if signal_type_filter == "OPTION_ADD" and option_add_kind_filter in ("new", "existing"):
        from app.wholesalers.models import Wholesaler
        ws_q = (
            db.session.query(
                Wholesaler.code, Wholesaler.name, db.func.count(ActionSignal.id).label("cnt")
            )
            .select_from(ActionSignal)
            .filter(ActionSignal.status == "pending", ActionSignal.signal_type == "OPTION_ADD")
            .join(MasterProduct, ActionSignal.master_product_id == MasterProduct.id)
            .join(Wholesaler, MasterProduct.wholesaler_id == Wholesaler.id)
            .join(StoreProduct, ActionSignal.store_product_id == StoreProduct.id)
        )
        if option_add_kind_filter == "new":
            ws_q = ws_q.filter(StoreProduct.applied_options_text.is_(None))
        else:
            ws_q = ws_q.filter(StoreProduct.applied_options_text.isnot(None))
        ws_q = ws_q.group_by(Wholesaler.code, Wholesaler.name).order_by(db.desc("cnt"))
        wholesaler_bulk_counts = [
            {"code": r.code, "name": r.name, "count": r.cnt} for r in ws_q.all()
        ]

    return render_template("actions.html", rows=rows, status_filter=status_filter,
                           pending_count=pending_count, failed_count=failed_count,
                           awaiting_input_count=awaiting_input_count,
                           tag_blocked_count=tag_blocked_count,
                           consumption_date_count=consumption_date_count,
                           close_status_count=close_status_count,
                           failure_kind_filter=failure_kind_filter,
                           pagination=pagination,
                           per_page=per_page, total=total,
                           all_stores=all_stores, store_filter=store_filter,
                           signal_type_filter=signal_type_filter,
                           option_type_filter=option_type_filter,
                           option_add_kind_filter=option_add_kind_filter,
                           opt_add_new_count=opt_add_new_count,
                           opt_add_existing_count=opt_add_existing_count,
                           wholesaler_bulk_counts=wholesaler_bulk_counts,
                           no_option_count=no_option_count,
                           option_no_extra_count=option_no_extra_count,
                           option_with_extra_count=option_with_extra_count,
                           search_q=search_q)


@actions_bp.route("/exclusions")
@login_required
def exclusions_page():
    from app.store.models import ProductExclusion
    exclusions = ProductExclusion.query.order_by(ProductExclusion.created_at.desc()).all()
    return render_template("exclusions.html", exclusions=exclusions)


@actions_bp.route("/exclusions/add", methods=["POST"])
@login_required
def add_exclusion():
    from app.store.models import StoreProduct, ProductExclusion
    store_product_id = request.json.get("store_product_id")
    reason = request.json.get("reason", "")
    store = StoreProduct.query.get_or_404(store_product_id)
    if store.exclusion:
        return jsonify({"ok": True})  # 이미 예외 등록됨
    db.session.add(ProductExclusion(store_product_id=store_product_id, reason=reason))
    # 기존 pending 시그널 스킵
    ActionSignal.query.filter_by(store_product_id=store_product_id, status="pending").update({"status": "skipped"})
    db.session.commit()
    return jsonify({"ok": True})


@actions_bp.route("/exclusions/<int:exclusion_id>/delete", methods=["POST"])
@login_required
def delete_exclusion(exclusion_id):
    from app.store.models import ProductExclusion
    exc = ProductExclusion.query.get_or_404(exclusion_id)
    db.session.delete(exc)
    db.session.commit()
    return jsonify({"ok": True})


@actions_bp.route("/actions/bulk-ids")
@login_required
def bulk_ids():
    """일괄 승인 대상 ids 조회 — 도매처별 OPTION_ADD 일괄 처리 전 ids 수집용."""
    from app.store.models import StoreProduct
    from app.wholesalers.models import Wholesaler

    signal_type = request.args.get("signal_type", "")
    option_add_kind = request.args.get("option_add_kind", "")
    wholesaler_code = request.args.get("wholesaler_code", "")
    status = request.args.get("status", "pending")

    if not signal_type or not wholesaler_code:
        return jsonify({"error": "signal_type and wholesaler_code required"}), 400

    ws = Wholesaler.query.filter_by(code=wholesaler_code).first()
    if not ws:
        return jsonify({"error": "wholesaler not found"}), 404

    query = (
        ActionSignal.query.filter_by(status=status, signal_type=signal_type)
        .join(MasterProduct, ActionSignal.master_product_id == MasterProduct.id)
        .filter(MasterProduct.wholesaler_id == ws.id)
    )
    if option_add_kind in ("new", "existing"):
        query = query.join(StoreProduct, ActionSignal.store_product_id == StoreProduct.id)
        if option_add_kind == "new":
            query = query.filter(StoreProduct.applied_options_text.is_(None))
        else:
            query = query.filter(StoreProduct.applied_options_text.isnot(None))

    ids = [r[0] for r in query.with_entities(ActionSignal.id).all()]

    # 대표 샘플 최대 5건 미리보기 생성 (detected_at 최신순)
    samples = []
    if ids:
        from app.settings import apply_margin, calculate_option_pricing
        sample_signals = (
            ActionSignal.query.filter(ActionSignal.id.in_(ids[:5]))
            .order_by(ActionSignal.detected_at.desc())
            .all()
        )
        for sig in sample_signals:
            try:
                m = sig.master
                s = sig.store
                if not m:
                    continue
                # 옵션 구조 계산
                if m.options_text and m.option_diffs:
                    opt = calculate_option_pricing(m.price, m.option_diffs)
                    list_price = opt["list_price"]
                    additions = opt["additions"]
                else:
                    list_price = apply_margin(m.price) if m.price else 0
                    additions = []
                names = [n.strip() for n in (m.options_text or "").split("\n") if n.strip()]
                new_options = []
                for i, name in enumerate(names):
                    add = additions[i] if i < len(additions) else 0
                    new_options.append({
                        "name": name,
                        "addition": add,
                        "price": list_price + add,
                    })
                samples.append({
                    "product_name": m.product_name or "-",
                    "seller_code": (s.seller_management_code if s else "") or "-",
                    "current_sale_price": (s.sale_price if s else None),
                    "wholesale_price": m.price,
                    "expected_list_price": list_price,
                    "new_options": new_options,
                })
            except Exception:
                # 샘플 하나 실패해도 나머지는 반환
                continue

    return jsonify({"ids": ids, "count": len(ids), "samples": samples})


@actions_bp.route("/actions/bulk-resolve", methods=["POST"])
@login_required
def bulk_resolve():
    import time
    ids = request.json.get("ids", [])
    action = request.json.get("action")  # approve / reject / skip

    ok_count = 0
    fail_count = 0
    awaiting_count = 0

    for signal_id in ids:
        signal = ActionSignal.query.get(signal_id)
        if not signal or signal.status != "pending":
            continue
        if action == "approve":
            _execute_signal(signal)
            if signal.status == "executed":
                ok_count += 1
            elif signal.status == "awaiting_input":
                awaiting_count += 1
            else:
                fail_count += 1
            time.sleep(0.3)  # Naver API rate limit 방지
        elif action == "reject":
            signal.status = "rejected"
            signal.resolved_at = kst_now()
            db.session.commit()
            ok_count += 1
        elif action == "skip":
            signal.status = "skipped"
            signal.resolved_at = kst_now()
            if signal.signal_type in ("OPTION_ADD", "OPTION_PRICE_CHANGE") and signal.store:
                _sugg = json.loads(signal.suggested_value or "{}")
                if _sugg.get("options_text") or _sugg.get("option_diffs"):
                    signal.store.applied_options_text = _sugg.get("options_text")
                    signal.store.applied_option_diffs = _sugg.get("option_diffs")
                    signal.store.applied_option_base_price = _sugg.get("base_price")
            db.session.commit()
            ok_count += 1

    return jsonify({"ok": True, "ok_count": ok_count, "fail_count": fail_count, "awaiting_count": awaiting_count})


@actions_bp.route("/actions/<int:signal_id>/resolve", methods=["POST"])
@login_required
def resolve_signal(signal_id):
    action = request.json.get("action")  # approve / reject / skip
    signal = ActionSignal.query.get_or_404(signal_id)

    try:
        if action == "approve":
            _execute_signal(signal)
            # 실행 후 실제 상태 확인 — 내부 오류로 failed가 됐어도 감지
            if signal.status == "awaiting_input":
                return jsonify({
                    "ok": False,
                    "status": "awaiting_input",
                    "required_fields": json.loads(signal.required_fields_missing or "[]"),
                }), 200
            if signal.status == "failed":
                return jsonify({"ok": False, "error": signal.error_message or "실행 실패"}), 200
        elif action == "reject":
            signal.status = "rejected"
            signal.resolved_at = kst_now()
            db.session.commit()
        elif action == "skip":
            signal.status = "skipped"
            signal.resolved_at = kst_now()
            # OPTION_ADD/OPTION_PRICE_CHANGE 건너뜀 → 현재 상태를 "적용됨"으로 기록
            if signal.signal_type in ("OPTION_ADD", "OPTION_PRICE_CHANGE") and signal.store:
                _sugg = json.loads(signal.suggested_value or "{}")
                if _sugg.get("options_text") or _sugg.get("option_diffs"):
                    signal.store.applied_options_text = _sugg.get("options_text")
                    signal.store.applied_option_diffs = _sugg.get("option_diffs")
                    signal.store.applied_option_base_price = _sugg.get("base_price")
            db.session.commit()
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

    return jsonify({"ok": True})


@actions_bp.route("/actions/bulk-retry", methods=["POST"])
@login_required
def bulk_retry():
    ids = request.json.get("ids", [])
    for signal_id in ids:
        signal = ActionSignal.query.get(signal_id)
        if signal and signal.status == "failed":
            signal.status = "pending"
            signal.error_message = None
            signal.resolved_at = None
    db.session.commit()
    return jsonify({"ok": True})


@actions_bp.route("/actions/<int:signal_id>/retry", methods=["POST"])
@login_required
def retry_signal(signal_id):
    signal = ActionSignal.query.get_or_404(signal_id)
    if signal.status != "failed":
        return jsonify({"ok": False, "error": "실패 상태 항목만 재시도할 수 있습니다."}), 400
    signal.status = "pending"
    signal.error_message = None
    signal.resolved_at = None
    db.session.commit()
    return jsonify({"ok": True})


def _is_tag_blocked_failure(signal: ActionSignal) -> bool:
    """sellerTags 등록불가 단어로 실패한 시그널인지 판별."""
    return (
        signal.status == "failed"
        and bool(signal.error_message)
        and "등록불가인 단어(" in signal.error_message
    )


def _force_tag_strip_retry(signal: ActionSignal) -> str:
    """ContextVar로 강제 모드 켜고 _execute_signal 재실행.
    반환값: signal.status ('executed' / 'failed' / 'awaiting_input')."""
    signal.status = "pending"
    signal.error_message = None
    signal.required_fields_missing = None
    signal.resolved_at = None
    db.session.commit()
    token = _force_strip_tags_ctx.set(True)
    try:
        _execute_signal(signal)
    finally:
        _force_strip_tags_ctx.reset(token)
    return signal.status


@actions_bp.route("/actions/<int:signal_id>/retry-tag-strip", methods=["POST"])
@login_required
def retry_signal_tag_strip(signal_id):
    """등록불가 단어로 실패한 시그널을 강제 모드로 재실행.
    sellerTags의 부분 일치 태그까지 모두 자동 제거 후 PUT 재시도."""
    signal = ActionSignal.query.get_or_404(signal_id)
    if not _is_tag_blocked_failure(signal):
        return jsonify({"ok": False, "error": "등록불가 태그 실패 항목만 처리할 수 있습니다."}), 400
    final_status = _force_tag_strip_retry(signal)
    return jsonify({
        "ok": final_status == "executed",
        "status": final_status,
        "error": signal.error_message if final_status != "executed" else None,
    })


@actions_bp.route("/actions/bulk-retry-tag-strip", methods=["POST"])
@login_required
def bulk_retry_tag_strip():
    """등록불가 단어 실패 시그널 일괄 강제 재시도. body: {"ids": [...]}"""
    import time
    ids = (request.json or {}).get("ids", []) or []
    ok_count = 0
    fail_count = 0
    skipped = 0
    for signal_id in ids:
        signal = ActionSignal.query.get(signal_id)
        if not signal or not _is_tag_blocked_failure(signal):
            skipped += 1
            continue
        try:
            final_status = _force_tag_strip_retry(signal)
            if final_status == "executed":
                ok_count += 1
            else:
                fail_count += 1
        except Exception:
            fail_count += 1
        time.sleep(0.3)  # Naver API rate limit 방지
    return jsonify({
        "ok": True,
        "ok_count": ok_count,
        "fail_count": fail_count,
        "skipped": skipped,
    })


@actions_bp.route("/actions/export/close-status.xlsx", methods=["GET"])
@login_required
def export_close_status_xlsx():
    """판매종료(CLOSE) 등 statusType 변경 불가로 failed된 시그널 일괄 엑셀 다운로드.
    Naver API change_status는 SALE/SUSPENSION/WAIT/SOLDOUT만 변경 가능 — CLOSE/PROHIBITION은
    어드민에서 수동 복구 필요. 운영자가 일괄 확인할 수 있게 export."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from datetime import datetime

    _close_sub_for_xlsx = (
        db.session.query(StoreProduct.id)
        .filter(StoreProduct.store_status.in_(("CLOSE", "PROHIBITION", "UNADMISSION")))
    )
    rows = (
        ActionSignal.query
        .filter_by(status="failed")
        .filter(db.or_(
            ActionSignal.error_message.like("%판매상태 변경이 가능합니다%"),
            ActionSignal.store_product_id.in_(_close_sub_for_xlsx),
        ))
        .order_by(ActionSignal.detected_at.desc())
        .all()
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "판매종료 복구필요"

    headers = [
        "signal_id",
        "signal_type",
        "도매처",
        "스토어",
        "상품명",
        "판매자관리코드",
        "원상품번호",
        "채널상품번호",
        "현재 판매상태",
        "감지일시",
        "오류 내용",
    ]
    ws.append(headers)
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="FFE5E7EB", end_color="FFE5E7EB", fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    for s in rows:
        store = s.store
        master = s.master
        wholesaler = master.wholesaler.name if master and master.wholesaler else ""
        store_name = store.naver_store.store_name if store and store.naver_store else ""
        ws.append([
            s.id,
            s.signal_type or "",
            wholesaler,
            store_name,
            master.product_name if master else (store.product_name if store else ""),
            store.seller_management_code if store else "",
            store.origin_product_no if store else "",
            store.channel_product_no if store else "",
            store.store_status if store else "",
            s.detected_at.strftime("%Y-%m-%d %H:%M") if s.detected_at else "",
            (s.error_message or "")[:500],
        ])

    widths = [10, 22, 14, 14, 50, 28, 14, 14, 14, 18, 80]
    for i, w in enumerate(widths, start=1):
        col = chr(64 + i) if i <= 26 else ("A" + chr(64 + i - 26))
        ws.column_dimensions[col].width = w

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)

    filename = f"actions_close_status_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        bio,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


@actions_bp.route("/actions/<int:signal_id>/dump-origin-notice", methods=["GET"])
@login_required
def dump_origin_notice(signal_id):
    """디버그 — 시그널의 store_product의 originProduct를 GET해서 productInfoProvidedNotice 부분을
    JSON으로 그대로 반환. 어드민에서 직접입력 모드로 처리해놓은 상품에 호출하면, 자유 텍스트가
    어떤 필드명에 들어가는지 정확히 보인다 (Naver의 비공개 스키마 역엔지니어링용)."""
    signal = ActionSignal.query.get_or_404(signal_id)
    store = signal.store
    if not store or not store.naver_store or not store.origin_product_no:
        return jsonify({"ok": False, "error": "store / origin_product_no 정보 없음"}), 400
    from store.naver.products import get_origin_product
    try:
        product_data = get_origin_product(
            store.origin_product_no,
            store.naver_store.client_id,
            store.naver_store.client_secret,
        )
    except Exception as e:
        return jsonify({"ok": False, "error": _parse_naver_error(e)}), 500
    origin = product_data.get("originProduct", {}) or {}
    detail = origin.get("detailAttribute", {}) or {}
    return jsonify({
        "ok": True,
        "origin_product_no": store.origin_product_no,
        "productInfoProvidedNoticeType": detail.get("productInfoProvidedNoticeType"),
        "productInfoProvidedNotice": detail.get("productInfoProvidedNotice"),
        # 단위가격도 함께 — 같은 디버그 흐름에서 정확한 키 확인
        "unitCapacity": detail.get("unitCapacity"),
    })


@actions_bp.route("/actions/<int:signal_id>/detail-html", methods=["GET"])
@login_required
def signal_detail_html(signal_id):
    """필수값 입력 모달에서 상품 정체 확인용 상세 HTML 미리보기.
    master.detail_description(도매처 수집본)을 sanitize 후 반환."""
    signal = ActionSignal.query.get_or_404(signal_id)
    raw = (signal.master.detail_description if signal.master else "") or ""
    if not str(raw).strip():
        return jsonify({"ok": False, "error": "도매처 상세 HTML이 비어 있습니다."})
    try:
        from store.naver.detail_html import sanitize_detail_html
        cleaned, _ = sanitize_detail_html(str(raw))
    except Exception:
        cleaned = str(raw)  # sanitize 실패 시 원본 폴백 (iframe sandbox로 격리됨)
    return jsonify({
        "ok": True,
        "html": cleaned,
        "raw_length": len(str(raw)),
    })


# 입력대기(awaiting_input) 흐름이 지원되는 시그널 타입 화이트리스트.
# _put_with_safe_retry를 거치는 모든 시그널이 ContextVar로 자동 적용되므로 추가 코드 변경 없이 확장 가능.
SUPPORTED_AWAITING_INPUT_TYPES = frozenset({
    "RESUME_POSSIBLE",
    "OPTION_ADD",
    "OPTION_PRICE_CHANGE",
    "OPTION_STOCK_CHANGE",
    "OPTION_STOCK_REFILL_NEEDED",
})


@actions_bp.route("/actions/<int:signal_id>/fill-required-fields", methods=["POST"])
@login_required
def fill_required_fields(signal_id):
    """awaiting_input 상태의 시그널에 운영자가 필수값을 채워 PUT 재시도.

    body: {"values": {"<dotted.field.path>": "<user_value>", ...}}

    동작:
      - 입력값을 ContextVar로 흘려 _execute_signal 재실행 → 모든 PUT에 자동 주입.
      - 성공: executed
      - KNOWN 필드 또 누락: required_fields_missing 갱신 후 awaiting_input 유지
      - 매치 없는 다른 에러: failed
    """
    from app.actions.required_fields import KNOWN_REQUIRED_FIELDS

    signal = ActionSignal.query.get_or_404(signal_id)
    if signal.status != "awaiting_input":
        return jsonify({"ok": False, "error": "입력 대기 상태 항목만 처리할 수 있습니다."}), 400
    if signal.signal_type not in SUPPORTED_AWAITING_INPUT_TYPES:
        return jsonify({"ok": False, "error": f"입력대기 지원 시그널이 아닙니다: {signal.signal_type}"}), 400

    body = request.get_json(silent=True) or {}
    raw_values = body.get("values") or {}
    if not isinstance(raw_values, dict) or not raw_values:
        return jsonify({"ok": False, "error": "입력값이 비어 있습니다."}), 400

    # 알려진 필드만 통과 + 빈 값은 페이로드에서 제외(보조 필드는 비울 수 있음).
    user_values: dict[str, str] = {}
    for k, v in raw_values.items():
        if k not in KNOWN_REQUIRED_FIELDS:
            return jsonify({"ok": False, "error": f"허용되지 않은 필드: {k}"}), 400
        if v is None or str(v).strip() == "":
            continue  # 빈 값은 PUT 본문에 안 넣음 — Naver가 알아서 처리
        user_values[k] = str(v).strip()
    if not user_values:
        return jsonify({"ok": False, "error": "최소 1개 이상의 값을 입력해야 합니다."}), 400

    _execute_signal_with_user_values(signal, user_values)

    if signal.status == "executed":
        return jsonify({"ok": True, "status": "executed"})
    if signal.status == "awaiting_input":
        return jsonify({
            "ok": False,
            "status": "awaiting_input",
            "required_fields": json.loads(signal.required_fields_missing or "[]"),
        })
    return jsonify({"ok": False, "status": "failed", "error": signal.error_message or "실행 실패"})


def _execute_signal_with_user_values(signal: ActionSignal, user_values: dict[str, str]):
    """fill_required_fields 라우트 전용 — 입력값을 ContextVar로 set하고 _execute_signal 재실행.
    시그널 타입(RESUME_POSSIBLE / OPTION_*)에 무관하게 디스패처가 알아서 분기.
    PUT 호출은 _put_with_safe_retry가 컨텍스트의 입력값을 자동 주입한다."""
    # 입력대기 → pending으로 되돌리고 _execute_signal이 정상 흐름 재실행하도록.
    signal.status = "pending"
    signal.error_message = None
    signal.required_fields_missing = None
    signal.resolved_at = None
    db.session.commit()

    token = _pending_user_values_ctx.set(dict(user_values))
    try:
        _execute_signal(signal)  # 결과 status/error_message/required_fields_missing은 디스패처가 갱신
    finally:
        _pending_user_values_ctx.reset(token)


@actions_bp.route("/actions/<int:signal_id>/revert", methods=["POST"])
@login_required
def revert_signal(signal_id):
    signal = ActionSignal.query.get_or_404(signal_id)
    if signal.status != "executed":
        return jsonify({"ok": False, "error": "실행된 항목만 되돌릴 수 있습니다."}), 400
    try:
        _revert_signal(signal)
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500
    return jsonify({"ok": True})


def _revert_signal(signal: ActionSignal):
    from store.naver import update_price, change_status

    store = signal.store
    current = json.loads(signal.current_value) if signal.current_value else {}

    if not store or not store.naver_store:
        raise ValueError("스토어 정보 없음")

    client_id = store.naver_store.client_id
    client_secret = store.naver_store.client_secret

    if signal.signal_type in ("PRICE_UP_NEEDED", "PRICE_DOWN_POSSIBLE"):
        orig_price = current.get("sale_price")
        if orig_price:
            update_price(store.origin_product_no, int(orig_price), client_id=client_id, client_secret=client_secret)
            store.sale_price = orig_price

    elif signal.signal_type in ("SUSPEND_NEEDED", "RESUME_POSSIBLE", "DISCONTINUE_NEEDED"):
        orig_status = current.get("store_status")
        if orig_status:
            change_status(store.origin_product_no, orig_status, client_id=client_id, client_secret=client_secret)
            store.store_status = orig_status

    elif signal.signal_type in ("OPTION_PRICE_CHANGE", "OPTION_STOCK_CHANGE"):
        raise ValueError("옵션 변동은 되돌리기를 지원하지 않습니다. 직접 수동으로 수정해주세요.")

    signal.status = "reverted"
    signal.resolved_at = kst_now()
    db.session.commit()


def _parse_naver_error(e) -> str:
    """Naver API HTTPError에서 사람이 읽을 수 있는 오류 메시지 추출.
    invalidInputs[].name(필드 경로)이 있으면 메시지 앞에 prefix로 붙여 진단 용이성 확보."""
    try:
        import requests as req_lib
        if isinstance(e, req_lib.HTTPError) and e.response is not None:
            data = e.response.json()
            invalid = data.get("invalidInputs") or []
            if invalid:
                parts = []
                for i in invalid:
                    msg = i.get("message", "")
                    if not msg:
                        continue
                    name = i.get("name", "")
                    parts.append(f"{name}: {msg}" if name else msg)
                if parts:
                    return " / ".join(parts)
            return data.get("message") or str(e)
    except Exception:
        pass
    return str(e)


# ---------------------------------------------------------------------------
# Naver PUT 자동 보강 (화이트리스트 — sellerTags 등록불가 단어만)
# ---------------------------------------------------------------------------

def _extract_naver_invalid_inputs(e) -> list:
    """Naver HTTPError 응답에서 invalidInputs 리스트만 추출.
    실패 시 빈 리스트 반환 (raise 안 함 — 호출부에서 안전 처리)."""
    try:
        import requests as req_lib
        if isinstance(e, req_lib.HTTPError) and e.response is not None:
            data = e.response.json()
            invalid = data.get("invalidInputs")
            if isinstance(invalid, list):
                return invalid
    except Exception:
        pass
    return []


def _extract_forbidden_tag_word(message: str):
    """'태그 항목에 등록불가인 단어(WORD)가 포함되어 있습니다' 메시지에서 단일 WORD 추출.
    Naver가 단어 여러 개를 콤마로 묶어 보내는 경우는 _extract_forbidden_tag_words 사용.
    매치 안 되면 None 반환. (하위 호환용 — 신규 코드는 복수형 사용)"""
    if not message:
        return None
    import re
    m = re.search(r"등록불가인 단어\(([^)]+)\)", message)
    return m.group(1).strip() if m else None


def _extract_forbidden_tag_words(message: str) -> list[str]:
    """'등록불가인 단어(W1,W2 ...)' 메시지에서 단어 리스트 추출.
    Naver가 단어 여러 개를 한 entry에 콤마/공백으로 묶어 보내는 케이스 대응.
    빈 토큰은 제거. 매치 없으면 빈 리스트.
    """
    if not message:
        return []
    import re
    m = re.search(r"등록불가인 단어\(([^)]+)\)", message)
    if not m:
        return []
    inner = m.group(1)
    # 콤마/공백/탭/슬래시(드물지만) 모두 분리자로 취급
    tokens = re.split(r"[,\s/]+", inner)
    return [t.strip() for t in tokens if t and t.strip()]


def _remove_seller_tags_containing(payload: dict, word: str) -> int:
    """payload의 sellerTags에서 word를 부분 문자열로 포함하는 모든 태그 제거.

    정확 일치는 물론 '강력양면테이프' ⊃ '양면테이프' 같은 부분 일치도 잡는다.
    안전장치 없는 강제 모드 — 강제 재시도 라우트에서만 사용.
    제거 개수 반환.
    """
    if not word:
        return 0
    word_norm = word.strip()
    if not word_norm:
        return 0
    origin = (payload or {}).get("originProduct")
    if not isinstance(origin, dict):
        return 0
    detail = origin.get("detailAttribute")
    if not isinstance(detail, dict):
        return 0
    seo = detail.get("seoInfo")
    if not isinstance(seo, dict):
        return 0
    tags = seo.get("sellerTags")
    if not isinstance(tags, list):
        return 0
    new_tags = []
    removed = 0
    for tag in tags:
        if isinstance(tag, dict):
            tag_text = (tag.get("text") or "").strip()
        else:
            tag_text = str(tag or "").strip()
        if word_norm in tag_text:  # 부분 일치 포함
            removed += 1
            continue
        new_tags.append(tag)
    if removed:
        seo["sellerTags"] = new_tags
    return removed


# 강제 모드 컨텍스트 — bulk_retry_tag_strip / retry_tag_strip 라우트에서만 True로 set.
# _put_with_safe_retry가 sellerTags 등록불가 단어 에러를 만나면 부분 일치까지 강제로 제거.
from contextvars import ContextVar
_force_strip_tags_ctx: ContextVar[bool] = ContextVar("force_strip_forbidden_tags", default=False)

# 입력대기(awaiting_input) → 사용자 입력값 컨텍스트.
# fill_required_fields 라우트가 _execute_signal 재실행 시 set해두면, 모든 PUT(_put_with_safe_retry)이
# 자동으로 payload에 입력값을 주입. 시그널 타입(RESUME_POSSIBLE / OPTION_*)에 무관하게 동작.
_pending_user_values_ctx: ContextVar[dict] = ContextVar("pending_user_values", default={})


def _remove_exact_seller_tag(payload: dict, target: str) -> bool:
    """payload의 originProduct.detailAttribute.seoInfo.sellerTags 에서
    `tag.strip() == target.strip()` 인 태그만 정확 일치로 제거.

    - 부분 문자열 매치 절대 금지 (예: target='보온' → '보온병' 보존)
    - 콤마/공백 분리 안 함 (태그 1개 = 한 덩어리 완성 문자열)
    - 태그가 dict({"text": "..."}) 형태든 string이든 둘 다 처리
    - 제거 발생 시 True, 아니면 False
    """
    if not target:
        return False
    target_norm = target.strip()
    if not target_norm:
        return False
    origin = (payload or {}).get("originProduct")
    if not isinstance(origin, dict):
        return False
    detail = origin.get("detailAttribute")
    if not isinstance(detail, dict):
        return False
    seo = detail.get("seoInfo")
    if not isinstance(seo, dict):
        return False
    tags = seo.get("sellerTags")
    if not isinstance(tags, list):
        return False
    new_tags = []
    removed = False
    for tag in tags:
        if isinstance(tag, dict):
            tag_text = (tag.get("text") or "").strip()
        else:
            tag_text = str(tag or "").strip()
        # 정확 일치만 제거
        if not removed and tag_text == target_norm:
            removed = True
            continue
        new_tags.append(tag)
    if removed:
        seo["sellerTags"] = new_tags
    return removed


class AwaitingInputNeeded(Exception):
    """Naver PUT이 KNOWN_REQUIRED_FIELDS 매치 누락으로 거부됐을 때 발생.
    _execute_signal의 except 블록이 잡아 status='awaiting_input'으로 보존한다."""
    def __init__(self, fields: list[dict]):
        super().__init__(f"awaiting input: {[f.get('name') for f in fields]}")
        self.fields = fields


def _put_with_safe_retry(origin_product_no, payload: dict, client_id: str, client_secret: str):
    """update_origin_product 호출 + 자동 보강 + 입력대기(awaiting_input) 분류.

    1) PUT 직전: _pending_user_values_ctx(awaiting_input → 사용자 입력값)이 있으면
       payload에 자동 주입 (fill_required_fields 라우트가 set한 컨텍스트).
       이로써 RESUME_POSSIBLE / OPTION_* 모든 시그널이 동일한 흐름으로 동작.

    2) sellerTags 등록불가 단어 자동 보강:
       - 단어 여러 개를 콤마/공백으로 묶어 보내는 케이스 포함, 각 단어를
         sellerTags 안에서 정확 일치로 제거 시도. (안전장치: 부분 일치 안 함)
       - 강제 모드(_force_strip_tags_ctx=True)일 때만 부분 일치까지 제거.
       하나라도 제거되면 1회 재시도.

    3) 자동 보강이 안 되는 경우 KNOWN_REQUIRED_FIELDS 매치 시도:
       - 매치 ≥ 1: AwaitingInputNeeded raise → 외부 except가 status='awaiting_input'로 보존.
       - 매치 0: 그대로 raise → 기존 failed 처리.
    """
    from store.naver.products import update_origin_product
    from app.actions.required_fields import (
        classify_invalid_inputs, apply_user_input, coerce_payload_booleans,
    )

    # (1) 입력대기 컨텍스트가 있으면 PUT 직전 주입 (Naver 타입 자동 변환 포함)
    user_values = _pending_user_values_ctx.get() or {}
    if user_values:
        apply_user_input(payload, user_values)

    # (1a) boolean 필드 방어 정규화 — 화이트리스트(NAVER_BOOLEAN_PATHS) 매치만 변환.
    # Yn 접미사 자동 변환은 절대 안 함(예: kcCertifiedProductExclusionYn은 enum 타입이라 거부됨).
    _bool_converted, _bool_paths = coerce_payload_booleans(payload)
    if _bool_converted:
        logger.info(f"[bool-coerce] origin_no={origin_product_no} converted={_bool_converted} paths={_bool_paths}")

    # (1b) statusType 안전 정규화 — Naver PUT은 시스템 관리값을 거부.
    # _execute_resume은 이미 SUSPENSION으로 강제하지만, OPTION_* 분기들은 GET 응답을 그대로 쓰므로
    # 여기서 한 번 더 안전하게 교체. 정상값(SALE/SUSPENSION/WAIT)은 그대로 유지.
    # PUT 거부 enum: SOLDOUT, CLOSE, PROHIBITION, UNADMISSION(미승인) — 모두 SUSPENSION으로 교체.
    _origin = (payload or {}).get("originProduct")
    if isinstance(_origin, dict):
        _st = _origin.get("statusType")
        if _st in ("SOLDOUT", "CLOSE", "PROHIBITION", "UNADMISSION") or not _st:
            logger.info(f"[statusType-coerce] origin_no={origin_product_no} {_st!r} -> SUSPENSION")
            _origin["statusType"] = "SUSPENSION"

    # (1c) 오늘출발 재고수량(todayStockQuantity) 정규화 — 이 값을 1로 강제.
    # Naver 검증: originProduct.stockQuantity ≥ deliveryInfo.todayStockQuantity.
    # 우리가 옵션 재고를 줄이면 합산 stockQuantity가 todayStockQuantity 이하로 떨어져 거부됨.
    # 다른 deliveryInfo 필드는 절대 건드리지 않고, todayStockQuantity 키가 존재할 때만 1로 설정.
    if isinstance(_origin, dict):
        _di = _origin.get("deliveryInfo")
        if isinstance(_di, dict) and "todayStockQuantity" in _di:
            _prev = _di.get("todayStockQuantity")
            if _prev != 1:
                _di["todayStockQuantity"] = 1
                logger.info(f"[todayStock-coerce] origin_no={origin_product_no} {_prev!r} -> 1")

    # 디버그: PUT 직전 페이로드 진단 dump
    try:
        import json as _json
        _origin = (payload or {}).get("originProduct") or {}
        _detail = _origin.get("detailAttribute") or {}
        _full = _json.dumps(payload, ensure_ascii=False)
        logger.info(
            f"[put] origin_no={origin_product_no} payload_len={len(_full)} "
            f"user_keys={list(user_values.keys())}\n"
            f"        unitCapacity={_detail.get('unitCapacity')}\n"
            f"        productInfoProvidedNotice={_json.dumps(_detail.get('productInfoProvidedNotice'), ensure_ascii=False)[:1500]}"
        )
        # column 위치 디버깅: 길이가 4000자 근처 미만이면 본문 일부 로그(최대 4KB)
        if len(_full) <= 8000:
            logger.info(f"[put-payload] origin_no={origin_product_no} body={_full}")
    except Exception:
        pass

    try:
        return update_origin_product(origin_product_no, payload, client_id, client_secret)
    except Exception as e:
        invalids = _extract_naver_invalid_inputs(e)
        if not invalids:
            raise
        # (2) sellerTags 자동 보강 시도
        force_mode = _force_strip_tags_ctx.get()
        modified = False
        for inv in invalids:
            name = (inv.get("name") or "").strip()
            msg = inv.get("message") or ""
            if name != "originProduct.detailAttribute.seoInfo.sellerTags":
                continue
            words = _extract_forbidden_tag_words(msg)
            if not words:
                continue
            for w in words:
                if force_mode:
                    removed = _remove_seller_tags_containing(payload, w)
                    if removed > 0:
                        modified = True
                        logger.info(
                            f"[autofix-force] sellerTags '{w}' 포함 태그 {removed}개 제거 "
                            f"(origin_no={origin_product_no})"
                        )
                else:
                    if _remove_exact_seller_tag(payload, w):
                        modified = True
                        logger.info(
                            f"[autofix] sellerTags 등록불가 단어 정확 일치 제거: '{w}' "
                            f"(origin_no={origin_product_no})"
                        )
        if modified:
            return update_origin_product(origin_product_no, payload, client_id, client_secret)

        # (3) KNOWN 필드 매치 — 운영자 입력 대기 흐름으로 보존
        matched = classify_invalid_inputs(invalids)
        # 진단 로그: invalidInputs 원본 전체를 떠서 정확한 키/메시지 단서 확보
        try:
            logger.warning(
                "[put_failed] origin_no=%s invalidInputs=%s matched=%s",
                origin_product_no,
                json.dumps(invalids, ensure_ascii=False),
                [m.get("name") for m in matched],
            )
        except Exception:
            pass
        if matched:
            raise AwaitingInputNeeded(matched)
        raise  # 자동 처리 불가능한 다른 에러 → failed


# ---------------------------------------------------------------------------
# 판매재개(RESUME_POSSIBLE) 실행 헬퍼
# ---------------------------------------------------------------------------

def _execute_resume(signal, master, store, suggested: dict,
                    client_id: str, client_secret: str):
    """RESUME_POSSIBLE 실행: 재고 0 → 999 보강 후 SALE 전환.

    Naver는 재고 0 상태에서 SALE 전환을 거부함("재고수량 항목을 입력해 주세요").
    도매처가 active로 돌아온 시점이므로 999로 보강해 재개 가능 상태로 만든다.

    운영자 입력값(awaiting_input)은 _pending_user_values_ctx 컨텍스트로 전달되며,
    _put_with_safe_retry가 PUT 직전에 자동으로 payload에 주입한다. KNOWN 필드 매치 시
    AwaitingInputNeeded raise → 외부 except가 status='awaiting_input'으로 보존.
    """
    from store.naver import change_status
    from store.naver.products import get_origin_product

    new_status = suggested.get("store_status")  # "SALE"
    if not new_status:
        raise ValueError("RESUME_POSSIBLE: store_status 정보 없음")

    product_data = get_origin_product(store.origin_product_no, client_id, client_secret)
    origin = product_data.get("originProduct", {})
    option_info = origin.get("detailAttribute", {}).get("optionInfo", {}) or {}
    combinations = option_info.get("optionCombinations", []) or []

    refilled = 0
    needs_update = False

    if _has_options(master):
        # 옵션 상품: usable=True 옵션의 재고 0만 999로
        for combo in combinations:
            if combo.get("usable", True) is False:
                continue
            if (combo.get("stockQuantity") or 0) == 0:
                combo["stockQuantity"] = 999
                refilled += 1
                needs_update = True
    else:
        # 단품: originProduct.stockQuantity + Naver 잔존 옵션 모두 보정
        if (origin.get("stockQuantity") or 0) == 0:
            origin["stockQuantity"] = 999
            refilled += 1
            needs_update = True
        for combo in combinations:
            if combo.get("usable", True) is False:
                continue
            if (combo.get("stockQuantity") or 0) == 0:
                combo["stockQuantity"] = 999
                refilled += 1
                needs_update = True

    # 운영자 입력값이 있으면 PUT을 무조건 실행(재고 보정이 없어도 입력값 반영 필요).
    must_put = needs_update or bool(_pending_user_values_ctx.get())

    if must_put:
        if combinations:
            option_info["optionCombinations"] = combinations
            origin.setdefault("detailAttribute", {})["optionInfo"] = option_info
        # Naver PUT은 originProduct.statusType을 필수로 요구하며 SOLDOUT 등 시스템 관리값은 거부함.
        # writable 값인 SUSPENSION으로 명시 → 검증 통과. 실제 SALE 전환은 아래 change_status가 담당.
        origin["statusType"] = "SUSPENSION"
        payload = {"originProduct": origin}
        logger.info(
            f"[actions][resume] PUT 직전 statusType={origin.get('statusType')!r}, "
            f"refilled={refilled} (store_id={store.id})"
        )
        # _put_with_safe_retry가 컨텍스트의 입력값 주입 + KNOWN 매치 시 AwaitingInputNeeded raise
        _put_with_safe_retry(store.origin_product_no, payload, client_id, client_secret)
        logger.info(f"[actions][resume] 재고 보강 {refilled}개 → 999 (store_id={store.id})")

    change_status(store.origin_product_no, new_status, client_id=client_id, client_secret=client_secret)
    store.store_status = new_status


# ---------------------------------------------------------------------------
# 옵션 유형 판별 헬퍼
# ---------------------------------------------------------------------------

def _has_options(master) -> bool:
    """도매처 마스터에 옵션이 있는 상품인지 판별"""
    return bool(master and master.options_text and master.options_text.strip())


def _has_extra_price(master) -> bool:
    """옵션 추가금(0 이외 차액)이 실제로 존재하는지 판별"""
    if not master or not master.option_diffs or not master.option_diffs.strip():
        return False
    try:
        return any(int(v.strip()) != 0 for v in master.option_diffs.split("\n") if v.strip())
    except ValueError:
        return False


def _normalize_diffs_for_compare(diffs: str | None) -> str | None:
    """비교용 정규화: 차액 전부 0이면 NULL 과 동일하게 취급.
    저장 구조는 건드리지 않고 비교 시에만 사용한다.
    master.option_diffs 는 이미 _normalize_diffs()(app/master/__init__.py)로 정규화돼 저장되므로
    applied/executed 이력값(과거에 "0\n0\n0" 형태로 저장됐을 수 있음) 쪽을 정규화해서 비교한다."""
    if not diffs or not diffs.strip():
        return None
    try:
        if all(int(v.strip()) == 0 for v in diffs.split("\n") if v.strip()):
            return None
    except ValueError:
        return diffs
    return diffs


# ---------------------------------------------------------------------------
# [유형 1] 옵션 없는 상품 — 가격 실행
# ---------------------------------------------------------------------------

def _execute_price_no_option(store, suggested: dict, client_id: str, client_secret: str):
    """
    옵션이 없는(또는 마스터에 옵션 정보 없는) 상품의 가격 변동 반영.
    도매처 기준: 마스터에 옵션 없음 → Naver 기존 옵션 추가금도 전부 0으로 초기화.
    Naver에 옵션 구조가 있어도 추가금을 제거하고 salePrice만 업데이트한다.
    """
    from store.naver import update_price
    from store.naver.products import get_origin_product, update_origin_product
    from app.settings import apply_margin

    wholesale_price = suggested.get("sale_price")
    if not wholesale_price or wholesale_price <= 0:
        raise ValueError("도매가 정보 없음")

    new_price = apply_margin(wholesale_price)

    # Naver 상품 조회: 기존 옵션 추가금 확인
    product_data = get_origin_product(store.origin_product_no, client_id, client_secret)
    origin = product_data.get("originProduct", {})
    option_info = origin.get("detailAttribute", {}).get("optionInfo", {})
    combinations = option_info.get("optionCombinations", [])

    if not combinations:
        # Naver 옵션 없음 → 단순 가격 업데이트
        update_price(store.origin_product_no, new_price, client_id=client_id, client_secret=client_secret)

    elif len(combinations) == 1:
        combo_price = combinations[0].get("price", 0)
        if combo_price == 0:
            # 옵션 1개, 추가금 0 → 사실상 단품 → 단순 가격 업데이트
            update_price(store.origin_product_no, new_price, client_id=client_id, client_secret=client_secret)
        else:
            # 옵션 1개, 추가금 있음 → 2단계: 추가금 0 초기화 → 가격 변경
            combinations[0]["price"] = 0
            option_info["optionCombinations"] = combinations
            origin.setdefault("detailAttribute", {})["optionInfo"] = option_info
            payload = {
                "originProduct": origin,
                "smartstoreChannelProduct": product_data.get("smartstoreChannelProduct", {}),
            }
            _put_with_safe_retry(store.origin_product_no, payload, client_id, client_secret)
            logger.info(f"[actions][no_option] Step1: 1개 combo 추가금 → 0 (store_id={store.id})")
            update_price(store.origin_product_no, new_price, client_id=client_id, client_secret=client_secret)
            store.option_list_price = new_price
            store.option_discount_amount = None

    else:
        # 옵션 여러 개 → 자동 처리 불가 → 단품↔옵션 불일치 페이지에서 수동 처리
        raise ValueError(
            f"Naver에 옵션 {len(combinations)}개 존재 — "
            f"'단품↔옵션 불일치' 페이지(/option-mismatch)에서 처리해주세요."
        )

    store.sale_price = new_price
    logger.info(f"[actions][no_option] 가격 반영: store_id={store.id}, price={new_price}")


# ---------------------------------------------------------------------------
# [유형 2] 옵션 있음·추가금 없음 — 가격 실행
# ---------------------------------------------------------------------------

def _execute_price_option_no_extra(store, master, suggested: dict, client_id: str, client_secret: str):
    """
    옵션은 있지만 옵션간 가격 차이가 없는 상품의 가격 변동 반영.
    모든 옵션 combination.price = 0, salePrice = new_price 로 업데이트한다.
    """
    from store.naver.products import get_origin_product, update_origin_product
    from app.settings import apply_margin

    wholesale_price = suggested.get("sale_price")
    if not wholesale_price or wholesale_price <= 0:
        raise ValueError("도매가 정보 없음")

    new_price = apply_margin(wholesale_price)

    product_data = get_origin_product(store.origin_product_no, client_id, client_secret)
    origin = product_data.get("originProduct", {})
    option_info = origin.get("detailAttribute", {}).get("optionInfo", {})
    combinations = option_info.get("optionCombinations", [])

    if not combinations:
        # 스토어에 옵션 없음 → 마스터 옵션으로 신규 생성
        option_names = [n.strip() for n in (master.options_text or "").split("\n") if n.strip()]
        if not option_names:
            raise ValueError("스토어/마스터 모두 옵션 없음 (option_no_extra 경로)")
        combinations = [
            {"optionName1": name, "price": 0, "stockQuantity": 999, "usable": True}
            for name in option_names
        ]
        # 옵션 구조 신규 초기화: 속성 그룹명 없으면 Naver가 optionCombinations를 무시함
        option_info["optionCombinationGroupNames"] = {"optionGroupName1": "옵션"}
        logger.info(
            f"[actions][option_no_extra] 스토어 옵션 없음 → 마스터 {len(combinations)}개 신규 생성: store_id={store.id}"
        )

    # master에 없는 옵션 제거 (옵션 구조 동기화)
    master_names = [n.strip() for n in (master.options_text or "").split("\n") if n.strip()]
    if master_names:
        filtered = [
            c for c in combinations
            if (c.get("optionName1") or c.get("optionName2") or "") in master_names
        ]
        if filtered:
            combinations = filtered
            logger.info(
                f"[actions][option_no_extra] 옵션 구조 동기화: "
                f"{len(combinations)}개 유지 (master 기준): store_id={store.id}"
            )

    for combo in combinations:
        combo["price"] = 0  # 추가금 없음 — 전부 0원

    option_info["optionCombinations"] = combinations
    origin.setdefault("detailAttribute", {})["optionInfo"] = option_info
    origin["salePrice"] = new_price
    origin["customerBenefit"] = {}  # 즉시할인 없음

    payload = {
        "originProduct": origin,
        "smartstoreChannelProduct": product_data.get("smartstoreChannelProduct", {}),
    }
    _put_with_safe_retry(store.origin_product_no, payload, client_id, client_secret)

    store.sale_price = new_price
    store.option_list_price = new_price
    store.option_discount_amount = None
    store.applied_options_text = master.options_text   # 적용된 옵션명 기록
    store.applied_option_diffs = None                  # 추가금 없음
    store.applied_option_base_price = wholesale_price  # 도매가 기록
    logger.info(f"[actions][option_no_extra] 가격 반영: store_id={store.id}, price={new_price}")


# ---------------------------------------------------------------------------
# [유형 3] 옵션 있음·추가금 있음 — 가격 실행
# ---------------------------------------------------------------------------

def _clamp_combo_price(price: int, list_price: int) -> int:
    """salePrice(정가) 기준 스마트스토어 옵션가 허용 범위로 클램핑"""
    if list_price < 2000:
        return max(0, min(price, list_price))
    elif list_price < 10000:
        return max(-(list_price // 2), min(price, list_price))
    else:
        return max(-(list_price // 2), min(price, list_price // 2))

def _execute_price_option_with_extra(store, master, suggested: dict, client_id: str, client_secret: str, signal: "ActionSignal"):
    """
    옵션도 있고 옵션간 추가금 차이도 있는 상품의 가격 변동 반영.
    정가 + 즉시할인 + 옵션추가금을 세트로 업데이트한다.
    addon 정책이 설정된 옵션은 supplementProductInfo 로 별도 반영.
    처리 후 같은 상품의 pending OPTION_PRICE_CHANGE 시그널을 자동 스킵한다.
    """
    from store.naver.products import get_origin_product, update_origin_product
    from app.settings import apply_margin, calculate_option_pricing
    try:
        from app.option_review import get_option_policies, build_supplement_payload, sync_addon_supplement_ids
    except ImportError:
        get_option_policies = lambda mid: {}
        build_supplement_payload = lambda *a, **kw: []
        sync_addon_supplement_ids = lambda *a, **kw: None

    wholesale_price = suggested.get("sale_price")
    if not wholesale_price or wholesale_price <= 0:
        raise ValueError("도매가 정보 없음")

    new_price = apply_margin(wholesale_price)
    pricing = calculate_option_pricing(wholesale_price, master.option_diffs)
    option_names = [n.strip() for n in master.options_text.split("\n") if n.strip()]
    policies = get_option_policies(master.id)  # {name: keep/addon/exclude}

    product_data = get_origin_product(store.origin_product_no, client_id, client_secret)
    origin = product_data.get("originProduct", {})
    detail = origin.setdefault("detailAttribute", {})
    option_info = detail.get("optionInfo", {})
    combinations = option_info.get("optionCombinations", [])

    if not combinations:
        # 스토어에 옵션 없음 → 마스터 옵션으로 신규 생성
        if not option_names:
            raise ValueError("스토어/마스터 모두 옵션 없음 (option_with_extra 경로)")
        combinations = [
            {
                "optionName1": name,
                "price": pricing["additions"][i] if i < len(pricing["additions"]) else 0,
                "stockQuantity": 999,
                "usable": True,
            }
            for i, name in enumerate(option_names)
        ]
        # 옵션 구조 신규 초기화: 속성 그룹명 없으면 Naver가 optionCombinations를 무시함
        option_info["optionCombinationGroupNames"] = {"optionGroupName1": "옵션"}
        logger.info(
            f"[actions][option_with_extra] 스토어 옵션 없음 → 마스터 {len(combinations)}개 신규 생성: store_id={store.id}"
        )

    # keep 정책 옵션만 combo 업데이트; addon/exclude 는 제외
    new_combos = []
    for i, combo in enumerate(combinations):
        name = combo.get("optionName1") or combo.get("optionName2") or ""
        decision = policies.get(name, "keep")
        if decision in ("addon", "exclude"):
            continue
        matched_idx = next((j for j, n in enumerate(option_names) if n == name), None)
        if matched_idx is None and i < len(pricing["additions"]):
            logger.warning(
                f"[actions][option_with_extra] 옵션명 매칭 실패 → 순서 폴백 "
                f"(store_product_id={store.id}, combo_idx={i}, name='{name}')"
            )
            matched_idx = i
        if matched_idx is not None and matched_idx < len(pricing["additions"]):
            combo["price"] = pricing["additions"][matched_idx]
            new_combos.append(combo)
        else:
            # 마스터에 없는 Naver 콤보 → 제외 (범위 오류 방지)
            logger.warning(
                f"[actions][option_with_extra] 마스터 미매칭 콤보 제외 "
                f"(store_product_id={store.id}, combo_idx={i}, name='{name}')"
            )

    if not new_combos:  # 안전장치: 모두 addon/exclude인 경우 전체 유지
        new_combos = combinations

    # 스마트스토어 옵션가 범위 초과 방지: 정가 기준 허용 범위로 클램핑
    lp = pricing["list_price"]
    for combo in new_combos:
        combo["price"] = _clamp_combo_price(combo["price"], lp)

    option_info["optionCombinations"] = new_combos
    detail["optionInfo"] = option_info

    # addon 옵션 → supplementProductInfo
    supplement_products = build_supplement_payload(master, wholesale_price, policies, product_data)
    if supplement_products:
        detail.setdefault("supplementProductInfo", {})["supplementProducts"] = supplement_products

    origin["salePrice"] = pricing["list_price"]
    if pricing["discount"] > 0:
        origin["customerBenefit"] = {
            "immediateDiscountPolicy": {
                "discountMethod": {"value": pricing["discount"], "unitType": "WON"}
            }
        }
    else:
        origin["customerBenefit"] = {}

    payload = {
        "originProduct": origin,
        "smartstoreChannelProduct": product_data.get("smartstoreChannelProduct", {}),
    }
    resp_data = _put_with_safe_retry(store.origin_product_no, payload, client_id, client_secret)

    # supplement ID 동기화 (신규 생성 시 Naver가 ID 부여)
    if supplement_products:
        has_new = any("id" not in s for s in supplement_products)
        if has_new or not isinstance(resp_data, dict) or not resp_data:
            fresh = get_origin_product(store.origin_product_no, client_id, client_secret)
            sync_addon_supplement_ids(master.id, fresh)
        else:
            sync_addon_supplement_ids(master.id, resp_data)

    store.sale_price = new_price
    store.option_list_price = pricing["list_price"]
    store.option_discount_amount = pricing["discount"] or None
    store.applied_options_text = master.options_text
    store.applied_option_diffs = master.option_diffs
    store.applied_option_base_price = wholesale_price
    logger.info(
        f"[actions][option_with_extra] 가격 반영: store_id={store.id}, "
        f"list_price={pricing['list_price']}, discount={pricing['discount']}, "
        f"addon_count={len(supplement_products)}"
    )

    # 옵션가도 함께 처리됐으므로 pending OPTION_PRICE_CHANGE 자동 스킵
    pending_opt = ActionSignal.query.filter_by(
        store_product_id=store.id,
        signal_type="OPTION_PRICE_CHANGE",
        status="pending",
    ).first()
    if pending_opt:
        pending_opt.status = "skipped"
        pending_opt.error_message = "가격변동 시그널 실행 시 옵션가도 함께 처리됨"
        pending_opt.resolved_at = kst_now()


# ---------------------------------------------------------------------------
# 시그널 실행 디스패처
# ---------------------------------------------------------------------------

def _execute_signal(signal: ActionSignal):
    from store.naver import change_status
    from app import log_buffer

    try:
        store = signal.store
        master = signal.master
        suggested = json.loads(signal.suggested_value) if signal.suggested_value else {}

        if not store or not store.naver_store:
            raise ValueError("스토어 정보 없음")

        _sc = store.seller_management_code or ""
        log_buffer.push(f"[액션] 실행: {signal.signal_type} | {_sc}")

        client_id = store.naver_store.client_id
        client_secret = store.naver_store.client_secret

        # ── 가격 변동: 옵션 유형별 독립 실행 ──────────────────────────────
        if signal.signal_type in ("PRICE_UP_NEEDED", "PRICE_DOWN_POSSIBLE"):
            if not _has_options(master):
                _execute_price_no_option(store, suggested, client_id, client_secret)
            elif not _has_extra_price(master):
                _execute_price_option_no_extra(store, master, suggested, client_id, client_secret)
            else:
                _execute_price_option_with_extra(store, master, suggested, client_id, client_secret, signal)

        # ── 판매재개: 재고 0 옵션/단품을 999로 보강 후 SALE 전환 ─────────
        elif signal.signal_type == "RESUME_POSSIBLE":
            _execute_resume(signal, master, store, suggested, client_id, client_secret)

        # ── 판매중지·단종: 상태만 전환 ───────────────────────────────────
        elif signal.signal_type in ("SUSPEND_NEEDED", "DISCONTINUE_NEEDED"):
            new_status = suggested.get("store_status")
            if new_status:
                change_status(store.origin_product_no, new_status, client_id=client_id, client_secret=client_secret)
                store.store_status = new_status
            # 판매중지·단종 실행 시 pending 옵션 시그널 즉시 삭제
            # (판매중지 상태에서는 옵션 재고/가격/구성 변동이 고객에게 무의미)
            ActionSignal.query.filter(
                ActionSignal.store_product_id == store.id,
                ActionSignal.status == "pending",
                ActionSignal.signal_type.in_(["OPTION_ADD", "OPTION_PRICE_CHANGE", "OPTION_STOCK_CHANGE"]),
            ).delete(synchronize_session=False)
            logger.info(f"[actions] 판매중지/단종 실행 → pending 옵션 시그널 정리 (store_id={store.id})")

        # ── 옵션 재고 변동: 옵션 있는 상품 전용 ──────────────────────────
        elif signal.signal_type == "OPTION_STOCK_CHANGE":
            from store.naver.products import get_origin_product, update_origin_product

            option_names = [n.strip() for n in suggested.get("options_text", "").split("\n") if n.strip()]
            option_stocks = []
            for s in suggested.get("option_stocks", "").split("\n"):
                try:
                    option_stocks.append(int(s.strip()))
                except ValueError:
                    option_stocks.append(999)

            product_data = get_origin_product(store.origin_product_no, client_id, client_secret)
            origin = product_data.get("originProduct", {})
            option_info = origin.get("detailAttribute", {}).get("optionInfo", {})
            combinations = option_info.get("optionCombinations", [])

            if not combinations:
                # Naver에 옵션 없음 → 도매처(master) 기준으로 옵션 신규 생성
                master_opt_names = [n.strip() for n in (master.options_text or "").split("\n") if n.strip()]
                if not master_opt_names:
                    raise ValueError("스토어/마스터 모두 옵션 없음")
                combinations = [
                    {
                        "optionName1": name,
                        "price": 0,
                        "stockQuantity": option_stocks[i] if i < len(option_stocks) else 999,
                        "usable": True,
                    }
                    for i, name in enumerate(master_opt_names)
                ]
                option_info["optionCombinationGroupNames"] = {"optionGroupName1": "옵션"}
                logger.info(
                    f"[actions][option_stock] Naver 옵션 없음 → 마스터 기준 "
                    f"{len(combinations)}개 신규 생성 (store_id={store.id})"
                )

            # 오너클랜 다차원 안전 매칭 시도 — 완전 일치하면 (n1, n2)로 정확 매칭
            from app.store import _match_ownerclan_dim
            ownerclan_match = _match_ownerclan_dim(master, combinations)
            if ownerclan_match:
                pair_to_idx = ownerclan_match["combo_index_by_pair"]
                pairs = ownerclan_match["master_pairs"]
                for master_idx, pair in enumerate(pairs):
                    if master_idx >= len(option_stocks):
                        continue
                    combo_idx = pair_to_idx[pair]
                    combinations[combo_idx]["stockQuantity"] = max(0, option_stocks[master_idx])
                logger.info(
                    f"[actions][option_stock] 오너클랜 다차원 매칭 적용 — {len(pairs)}개 옵션 (store_id={store.id})"
                )
            else:
                # 기존 1차원 매칭 (옵션명 직접 비교 + 순서 폴백)
                for i, combo in enumerate(combinations):
                    name = combo.get("optionName1") or combo.get("optionName2") or ""
                    matched_idx = next((j for j, n in enumerate(option_names) if n == name), None)
                    if matched_idx is None and i < len(option_stocks):
                        logger.warning(
                            f"[actions][option_stock] 옵션명 매칭 실패 → 순서 폴백 "
                            f"(store_product_id={store.id}, combo_idx={i}, name='{name}')"
                        )
                        matched_idx = i
                    if matched_idx is not None and matched_idx < len(option_stocks):
                        combo["stockQuantity"] = max(0, option_stocks[matched_idx])

            option_info["optionCombinations"] = combinations
            origin.setdefault("detailAttribute", {})["optionInfo"] = option_info
            _put_with_safe_retry(store.origin_product_no, {"originProduct": origin}, client_id, client_secret)

        # ── SALE 상태에서 일부 옵션만 재고 0인 케이스 보강 ─────────────────
        # 도매처는 active인데 Naver 옵션만 0으로 남은 경우 — 999로 채워 손님이 다시 살 수 있게 함.
        # statusType은 이미 SALE이라 round-trip OK (RESUME_POSSIBLE의 SUSPENSION 트릭 불필요).
        elif signal.signal_type == "OPTION_STOCK_REFILL_NEEDED":
            from store.naver.products import get_origin_product, update_origin_product

            product_data = get_origin_product(store.origin_product_no, client_id, client_secret)
            origin = product_data.get("originProduct", {})
            option_info = origin.get("detailAttribute", {}).get("optionInfo", {}) or {}
            combinations = option_info.get("optionCombinations", []) or []

            if not combinations:
                raise ValueError("Naver 옵션 없음 — 보강 대상 아님")

            refilled = 0
            for combo in combinations:
                if combo.get("usable", True) is False:
                    continue
                if (combo.get("stockQuantity") or 0) == 0:
                    combo["stockQuantity"] = 999
                    refilled += 1

            if refilled == 0:
                raise ValueError("usable=True 옵션 중 재고 0인 옵션이 이미 없음 (다른 경로로 복구됨)")

            option_info["optionCombinations"] = combinations
            origin.setdefault("detailAttribute", {})["optionInfo"] = option_info
            _put_with_safe_retry(store.origin_product_no, {"originProduct": origin}, client_id, client_secret)
            logger.info(f"[actions][stock_refill] 옵션 재고 보강 {refilled}개 → 999 (store_id={store.id})")

        # ── 옵션 추가금 변동: 추가금 있는 상품 전용 ──────────────────────
        elif signal.signal_type == "OPTION_PRICE_CHANGE":
            from store.naver.products import get_origin_product, update_origin_product
            try:
                from app.option_review import get_option_policies, build_supplement_payload, sync_addon_supplement_ids
            except ImportError:
                get_option_policies = lambda mid: {}
                build_supplement_payload = lambda *a, **kw: []
                sync_addon_supplement_ids = lambda *a, **kw: None

            list_price   = suggested.get("list_price")
            discount     = suggested.get("discount", 0)
            additions    = suggested.get("additions", [])
            option_names = [n.strip() for n in suggested.get("options_text", "").split("\n") if n.strip()]
            base_price   = suggested.get("base_price")

            # 구형 시그널 폴백
            if not list_price or not additions:
                from app.settings import calculate_option_pricing
                _diffs = suggested.get("option_diffs", "")
                if not base_price or not _diffs:
                    raise ValueError("옵션 가격 데이터 부족 (base_price/option_diffs 없음)")
                _p = calculate_option_pricing(base_price, _diffs)
                list_price = _p["list_price"]
                discount   = _p["discount"]
                additions  = _p["additions"]

            policies = get_option_policies(master.id)

            product_data = get_origin_product(store.origin_product_no, client_id, client_secret)
            origin = product_data.get("originProduct", {})
            detail = origin.setdefault("detailAttribute", {})
            option_info = detail.get("optionInfo", {})
            combinations = option_info.get("optionCombinations", [])

            if not combinations:
                # Naver에 옵션 없음 → 도매처(master) 기준으로 옵션 신규 생성
                master_opt_names = [n.strip() for n in (master.options_text or "").split("\n") if n.strip()]
                if not master_opt_names:
                    raise ValueError("스토어/마스터 모두 옵션 없음")
                combinations = [
                    {
                        "optionName1": name,
                        "price": additions[i] if i < len(additions) else 0,
                        "stockQuantity": 999,
                        "usable": True,
                    }
                    for i, name in enumerate(master_opt_names)
                ]
                option_info["optionCombinationGroupNames"] = {"optionGroupName1": "옵션"}
                logger.info(
                    f"[actions][option_price] Naver 옵션 없음 → 마스터 기준 "
                    f"{len(combinations)}개 신규 생성 (store_id={store.id})"
                )

            # 오너클랜 다차원 안전 매칭 시도
            from app.store import _match_ownerclan_dim
            ownerclan_match = _match_ownerclan_dim(master, combinations)
            if ownerclan_match:
                pair_to_idx = ownerclan_match["combo_index_by_pair"]
                pairs = ownerclan_match["master_pairs"]
                # master 1차원 옵션명 (정책 키 — keep/addon/exclude 판정용)
                m_lines = [l.strip() for l in (master.options_text or "").split("\n") if l.strip()]
                new_combos = []
                for master_idx, pair in enumerate(pairs):
                    policy_key = m_lines[master_idx] if master_idx < len(m_lines) else ""
                    decision = policies.get(policy_key, "keep")
                    if decision in ("addon", "exclude"):
                        continue
                    combo_idx = pair_to_idx[pair]
                    combo = combinations[combo_idx]
                    if master_idx < len(additions):
                        combo["price"] = additions[master_idx]
                    new_combos.append(combo)
                if not new_combos:
                    new_combos = combinations
                logger.info(
                    f"[actions][option_price] 오너클랜 다차원 매칭 적용 — {len(pairs)}개 옵션 (store_id={store.id})"
                )
            else:
                # 기존 1차원 매칭
                new_combos = []
                for i, combo in enumerate(combinations):
                    name = combo.get("optionName1") or combo.get("optionName2") or ""
                    decision = policies.get(name, "keep")
                    if decision in ("addon", "exclude"):
                        continue
                    matched_idx = next((j for j, n in enumerate(option_names) if n == name), None)
                    if matched_idx is None and i < len(additions):
                        logger.warning(
                            f"[actions][option_price] 옵션명 매칭 실패 → 순서 폴백 "
                            f"(store_product_id={store.id}, combo_idx={i}, name='{name}')"
                        )
                        matched_idx = i
                    if matched_idx is not None and matched_idx < len(additions):
                        combo["price"] = additions[matched_idx]
                    new_combos.append(combo)

                if not new_combos:
                    new_combos = combinations

            option_info["optionCombinations"] = new_combos
            detail["optionInfo"] = option_info

            # addon 옵션 → supplementProductInfo
            wholesale_base = base_price or list_price
            supplement_products = build_supplement_payload(master, wholesale_base, policies, product_data)
            if supplement_products:
                detail.setdefault("supplementProductInfo", {})["supplementProducts"] = supplement_products

            origin["salePrice"] = list_price
            if discount > 0:
                origin["customerBenefit"] = {
                    "immediateDiscountPolicy": {
                        "discountMethod": {"value": discount, "unitType": "WON"}
                    }
                }
            else:
                origin["customerBenefit"] = {}

            payload = {
                "originProduct": origin,
                "smartstoreChannelProduct": product_data.get("smartstoreChannelProduct", {}),
            }
            resp_data = _put_with_safe_retry(store.origin_product_no, payload, client_id, client_secret)

            if supplement_products:
                has_new = any("id" not in s for s in supplement_products)
                if has_new or not isinstance(resp_data, dict) or not resp_data:
                    fresh = get_origin_product(store.origin_product_no, client_id, client_secret)
                    sync_addon_supplement_ids(master.id, fresh)
                else:
                    sync_addon_supplement_ids(master.id, resp_data)

            store.sale_price = list_price - discount
            store.option_list_price = list_price
            store.option_discount_amount = discount if discount > 0 else None
            store.applied_option_diffs = suggested.get("option_diffs") or None
            store.applied_option_base_price = suggested.get("base_price")
            store.applied_options_text = suggested.get("options_text") or master.options_text

        # ── 옵션 구성 전체 교체: 추가금 있는 상품 전용 ───────────────────
        elif signal.signal_type == "OPTION_ADD":
            from store.naver.products import get_origin_product, update_origin_product
            from app.settings import calculate_option_pricing
            try:
                from app.option_review import get_option_policies, build_supplement_payload, sync_addon_supplement_ids
            except ImportError:
                get_option_policies = lambda mid: {}
                build_supplement_payload = lambda *a, **kw: []
                sync_addon_supplement_ids = lambda *a, **kw: None

            base_price   = suggested.get("base_price") or (master.price if master else None)
            option_diffs = suggested.get("option_diffs") or (master.option_diffs if master else None) or ""
            options_text = suggested.get("options_text") or (master.options_text if master else "") or ""

            if not base_price or not options_text:
                raise ValueError("OPTION_ADD: base_price 또는 options_text 없음")

            master_names = [n.strip() for n in options_text.split("\n") if n.strip()]
            # option_diffs 없으면 추가금 없음(전부 0)으로 처리
            if not option_diffs:
                option_diffs = "\n".join(["0"] * len(master_names))
            pricing = calculate_option_pricing(base_price, option_diffs)
            additions = pricing["additions"]
            policies = get_option_policies(master.id)  # {name: keep/addon/exclude}

            product_data = get_origin_product(store.origin_product_no, client_id, client_secret)
            origin = product_data.get("originProduct", {})
            detail = origin.setdefault("detailAttribute", {})
            option_info = detail.get("optionInfo", {})

            # 기존 Naver 콤보를 기반으로 매칭 후 가격 업데이트
            # (2차원 옵션 구조 보존 — optionName1/optionName2 그대로 유지)
            existing_combos = option_info.get("optionCombinations", [])

            # master.option_stocks 파싱 (다차원 PUT 시 옵션별 재고 반영)
            master_stock_lines = [s.strip() for s in (master.option_stocks or "").split("\n") if s.strip()]
            def _stock_for(idx: int) -> int:
                if idx < len(master_stock_lines):
                    try:
                        return max(0, int(master_stock_lines[idx]))
                    except ValueError:
                        pass
                return 999

            # 오너클랜 다차원 안전 매칭 시도 — 완전 일치하면 (n1, n2)로 다차원 PUT
            from app.store import _match_ownerclan_dim, _is_ownerclan, _split_ownerclan_options
            ownerclan_match = _match_ownerclan_dim(master, existing_combos)

            # 오너클랜이고 master 모든 줄이 '_' 1개로 깨끗이 분리되면 다차원 빌드 자격.
            # _match_ownerclan_dim의 set equality가 안 맞아도(=master에 신규 옵션 추가됨)
            # master 기준 다차원 강제 빌드. 다른 도매처는 _is_ownerclan False로 영향 없음.
            ownerclan_master_pairs = (
                _split_ownerclan_options(master.options_text or "") if _is_ownerclan(master) else None
            )

            if ownerclan_match:
                pair_to_idx = ownerclan_match["combo_index_by_pair"]
                pairs = ownerclan_match["master_pairs"]
                new_combos = []
                for master_idx, pair in enumerate(pairs):
                    policy_key = master_names[master_idx] if master_idx < len(master_names) else ""
                    decision = policies.get(policy_key, "keep")
                    if decision in ("addon", "exclude"):
                        continue
                    combo_idx = pair_to_idx[pair]
                    combo = existing_combos[combo_idx]
                    if master_idx < len(additions):
                        combo["price"] = additions[master_idx]
                    combo["stockQuantity"] = _stock_for(master_idx)
                    new_combos.append(combo)
                # 기존 그룹명 보존 (이미 다차원이라 들어있을 것). 안전 가드만.
                groups = option_info.get("optionCombinationGroupNames") or {}
                if not groups.get("optionGroupName1") or not groups.get("optionGroupName2"):
                    groups = {"optionGroupName1": groups.get("optionGroupName1") or "옵션1",
                              "optionGroupName2": groups.get("optionGroupName2") or "옵션2"}
                    option_info["optionCombinationGroupNames"] = groups
                logger.info(
                    f"[actions][option_add] 오너클랜 다차원 매칭 적용 — {len(new_combos)}개 콤보 (store_id={store.id})"
                )
            elif ownerclan_master_pairs:
                # 오너클랜 + master 모든 줄 '_' 1개 → master 기준 다차원 강제 빌드.
                # Naver 기존 콤보의 optionName1/2 차원과 master의 차원이 동일하다는 전제하에 동작.
                # 기존 그룹명 보존하되, 1차원 등록된 상품에서 다차원으로 강제 변환은 안 함.
                existing_groups = option_info.get("optionCombinationGroupNames") or {}
                has_dim2_group = bool(existing_groups.get("optionGroupName2"))
                if has_dim2_group:
                    new_combos = []
                    for i, (n1, n2) in enumerate(ownerclan_master_pairs):
                        policy_key = master_names[i] if i < len(master_names) else ""
                        decision = policies.get(policy_key, "keep")
                        if decision in ("addon", "exclude"):
                            continue
                        new_combos.append({
                            "optionName1": n1,
                            "optionName2": n2,
                            "price": additions[i] if i < len(additions) else 0,
                            "stockQuantity": _stock_for(i),
                            "usable": True,
                        })
                    logger.info(
                        f"[actions][option_add] 오너클랜 master 기준 다차원 강제 빌드 — "
                        f"{len(new_combos)}개 콤보 (store_id={store.id})"
                    )
                else:
                    # Naver가 1차원 등록 상태면 강제 다차원 변환 위험 → 기존 1차원 흐름으로 폴백
                    ownerclan_master_pairs = None  # 아래 1차원 분기로 떨어뜨림

            if not ownerclan_match and not (ownerclan_master_pairs):
                # 기존 1차원 매칭 (옵션명 직접 비교 + 마스터에 없는 옵션 제거 + 신규 추가)
                new_combos = []
                handled_master_idx = set()

                for combo in existing_combos:
                    n1 = combo.get("optionName1") or ""
                    n2 = combo.get("optionName2") or ""
                    # 마스터 옵션명이 optionName1 또는 optionName2 중 하나와 일치하면 매칭
                    matched_idx = next(
                        (i for i, n in enumerate(master_names) if n == n1 or n == n2),
                        None
                    )
                    if matched_idx is None:
                        continue  # 마스터에 없는 옵션 → 제거
                    decision = policies.get(master_names[matched_idx], "keep")
                    if decision in ("addon", "exclude"):
                        continue
                    combo["price"] = additions[matched_idx] if matched_idx < len(additions) else 0
                    combo["stockQuantity"] = _stock_for(matched_idx)
                    new_combos.append(combo)
                    handled_master_idx.add(matched_idx)

                # 기존 Naver에 없는 신규 마스터 옵션 추가
                for i, name in enumerate(master_names):
                    if i in handled_master_idx:
                        continue
                    decision = policies.get(name, "keep")
                    if decision in ("addon", "exclude"):
                        continue
                    new_combos.append({
                        "optionName1":  name,
                        "price":        additions[i] if i < len(additions) else 0,
                        "stockQuantity": _stock_for(i),
                        "usable":       True,
                    })

                if not new_combos:  # 안전장치: master 기준 강제 생성 (기존 Naver 콤보 복원 금지)
                    logger.warning(
                        f"[actions][option_add] 매칭된 콤보 없음 → master 기준 강제 생성: store_id={store.id}"
                    )
                    new_combos = [
                        {"optionName1": n, "price": additions[i] if i < len(additions) else 0,
                         "stockQuantity": _stock_for(i), "usable": True}
                        for i, n in enumerate(master_names)
                    ]

            logger.info(
                f"[actions][option_add] 전체교체: store_id={store.id}, "
                f"combo={len(new_combos)}개 (addon/exclude 제외)"
            )

            option_info["optionCombinations"] = new_combos
            detail["optionInfo"] = option_info

            # addon 옵션 → supplementProductInfo
            supplement_products = build_supplement_payload(master, base_price, policies, product_data)
            if supplement_products:
                detail.setdefault("supplementProductInfo", {})["supplementProducts"] = supplement_products

            origin["salePrice"] = pricing["list_price"]
            if pricing["discount"] > 0:
                origin["customerBenefit"] = {
                    "immediateDiscountPolicy": {
                        "discountMethod": {"value": pricing["discount"], "unitType": "WON"}
                    }
                }
            else:
                origin["customerBenefit"] = {}

            payload = {
                "originProduct": origin,
                "smartstoreChannelProduct": product_data.get("smartstoreChannelProduct", {}),
            }
            resp_data = _put_with_safe_retry(store.origin_product_no, payload, client_id, client_secret)

            if supplement_products:
                has_new = any("id" not in s for s in supplement_products)
                if has_new or not isinstance(resp_data, dict) or not resp_data:
                    fresh = get_origin_product(store.origin_product_no, client_id, client_secret)
                    sync_addon_supplement_ids(master.id, fresh)
                else:
                    sync_addon_supplement_ids(master.id, resp_data)

            store.sale_price = pricing["sale_price"]
            store.option_list_price = pricing["list_price"]
            store.option_discount_amount = pricing["discount"] if pricing["discount"] > 0 else None
            store.applied_options_text = options_text
            store.applied_option_diffs = master.option_diffs  # 정규화된 master 값 (추가금 없으면 None)
            store.applied_option_base_price = base_price

            # 옵션 구조+가격 함께 처리됐으므로 pending OPTION_PRICE_CHANGE 자동 스킵
            pending_price = ActionSignal.query.filter_by(
                store_product_id=store.id,
                signal_type="OPTION_PRICE_CHANGE",
                status="pending",
            ).first()
            if pending_price:
                pending_price.status = "skipped"
                pending_price.error_message = "OPTION_ADD 실행 시 옵션 가격도 함께 처리됨"
                pending_price.resolved_at = kst_now()

            # OPTION_ADD PUT 성공으로 master와 네이버 옵션이 동기화됨 →
            # 이 상품의 옵션 불일치(pending)도 자동 해소. 화면 옛 캐시 잔상 방지.
            try:
                from app.store.models import StoreOptionMismatch
                for _mm in StoreOptionMismatch.query.filter_by(
                    store_product_id=store.id, status="pending"
                ).all():
                    _mm.status = "resolved"
                    _mm.resolved_at = kst_now()
            except Exception as _e:
                logger.debug(f"[option_add] mismatch 자동 해소 실패(무시): {_e}")

        # ── 상세페이지 갱신: 도매처 detail_description → 네이버 detailContent ──
        elif signal.signal_type == "DETAIL_CHANGE":
            from store.naver.products import sync_detail_content
            raw_html = (master.detail_description or "") if master else ""
            if not raw_html.strip():
                raise ValueError("DETAIL_CHANGE: master에 상세페이지 HTML 없음")
            sync_detail_content(
                store.origin_product_no,
                raw_html,
                client_id,
                client_secret,
            )

        signal.status = "executed"
        signal.error_message = None
        signal.required_fields_missing = None
        signal.resolved_at = kst_now()
        db.session.commit()
        log_buffer.push(f"[액션] 완료: {signal.signal_type} | {_sc}")

    except AwaitingInputNeeded as ai:
        # KNOWN_REQUIRED_FIELDS 매치된 PUT 거부 — 운영자 입력 대기 상태로 보존
        db.session.rollback()
        signal.status = "awaiting_input"
        signal.required_fields_missing = json.dumps(ai.fields, ensure_ascii=False)
        signal.error_message = None
        signal.resolved_at = None
        db.session.commit()
        _missing = ", ".join(f.get("label", f.get("name", "?")) for f in ai.fields)
        log_buffer.push(f"[액션] 입력대기: {signal.signal_type} | {_sc} | 누락: {_missing}")

    except Exception as e:
        db.session.rollback()
        signal.status = "failed"
        signal.error_message = _parse_naver_error(e)
        signal.resolved_at = kst_now()
        db.session.commit()
        log_buffer.push(f"[액션] 실패: {signal.signal_type} | {_sc} | {signal.error_message}")


def detect_action_signals(wholesaler_id: int) -> dict:
    """
    마스터 상품 vs 스토어 상품 비교 → ActionSignal 생성
    매 실행마다 기존 pending 시그널을 지우고 현재 상태로 새로 감지 (중복/충돌 방지)
    """
    stats = {
        "PRICE_UP_NEEDED": 0,
        "PRICE_DOWN_POSSIBLE": 0,
        "SUSPEND_NEEDED": 0,
        "RESUME_POSSIBLE": 0,
        "DISCONTINUE_NEEDED": 0,
        "OPTION_PRICE_CHANGE": 0,
        "OPTION_STOCK_CHANGE": 0,
        "OPTION_ADD": 0,
        "OPTION_STOCK_REFILL_NEEDED": 0,
    }

    # 해당 도매처의 매칭된 스토어 상품만 조회 — 전체 로드 방지, 관계 미리 로드
    stores = (
        StoreProduct.query
        .filter(StoreProduct.master_product_id.isnot(None))
        .join(MasterProduct, StoreProduct.master_product_id == MasterProduct.id)
        .filter(MasterProduct.wholesaler_id == wholesaler_id)
        .options(joinedload(StoreProduct.master), joinedload(StoreProduct.exclusion))
        .all()
    )

    store_ids = [s.id for s in stores]

    CHUNK = 500

    # 옵션 시그널(OPTION_ADD/OPTION_PRICE_CHANGE/OPTION_STOCK_CHANGE)은 값이 바뀔 때만 갱신.
    # pending을 먼저 로드한 뒤 PRICE/STATUS 시그널만 삭제한다.
    OPTION_TYPES = {"OPTION_ADD", "OPTION_PRICE_CHANGE", "OPTION_STOCK_CHANGE", "OPTION_STOCK_REFILL_NEEDED"}
    PRICE_STATUS_TYPES = [
        "PRICE_UP_NEEDED", "PRICE_DOWN_POSSIBLE",
        "SUSPEND_NEEDED", "RESUME_POSSIBLE", "DISCONTINUE_NEEDED",
    ]

    # 기존 pending 옵션 시그널 로드 (master_id, store_id, type) → ActionSignal 객체
    prev_opts: dict[tuple, ActionSignal] = {}
    for i in range(0, len(store_ids), CHUNK):
        chunk = store_ids[i:i + CHUNK]
        for sig in ActionSignal.query.filter(
            ActionSignal.store_product_id.in_(chunk),
            ActionSignal.status == "pending",
            ActionSignal.signal_type.in_(list(OPTION_TYPES)),
        ).all():
            prev_opts[(sig.master_product_id, sig.store_product_id, sig.signal_type)] = sig

    # PRICE/STATUS pending만 삭제 → 현재 상태로 재생성
    for i in range(0, len(store_ids), CHUNK):
        chunk = store_ids[i:i + CHUNK]
        # PRICE/STATUS 신호 전체 삭제 (pending + executed + reverted + failed 등)
        # → 현재 수집 결과 기준으로 새로 생성 (항상 최신 1건만 유지)
        ActionSignal.query.filter(
            ActionSignal.store_product_id.in_(chunk),
            ActionSignal.signal_type.in_(PRICE_STATUS_TYPES),
        ).delete(synchronize_session=False)
    if store_ids:
        db.session.flush()

    # 현재 루프에서 이미 처리된 시그널 추적 (기존 pending 옵션 시그널 포함)
    existing_pending = set(prev_opts.keys())

    for store in stores:
        master = store.master

        if not master:
            continue

        if store.exclusion:
            continue

        # 판매중지·판매종료 상품은 상태 시그널만 처리 (가격/옵션 시그널 스킵)
        if store.store_status != "SALE":
            # 기존 pending 옵션 시그널 정리 (판매중지 상품에는 불필요)
            for otype in OPTION_TYPES:
                okey = (master.id, store.id, otype)
                if okey in prev_opts:
                    db.session.delete(prev_opts.pop(okey))
                    existing_pending.discard(okey)
            _check_status_signals(master, store, stats, existing_pending)
            continue

        _check_price_signals(master, store, stats, existing_pending)
        _check_status_signals(master, store, stats, existing_pending)
        _check_option_add_signals(master, store, stats, existing_pending, prev_opts)
        _check_option_signals(master, store, stats, existing_pending, prev_opts)
        _check_option_stock_signals(master, store, stats, existing_pending, prev_opts)
        _check_option_stock_refill_signals(master, store, stats, existing_pending, prev_opts)

    db.session.commit()
    logger.info(f"[actions] 시그널 감지 완료: {stats}")
    return stats


def _check_price_signals(master: MasterProduct, store: StoreProduct, stats: dict, pending: set):
    if not master.price or not store.sale_price:
        return

    from app.settings import apply_margin, calculate_option_pricing

    # 기대 정가 계산 (옵션 상품은 가장 싼 옵션 기준, 단품은 도매가 기준)
    if master.options_text and master.option_diffs:
        try:
            margin_price = calculate_option_pricing(master.price, master.option_diffs)["list_price"]
        except Exception:
            # 옵션 파싱 실패 시 단품 계산으로 fallback
            margin_price = apply_margin(master.price)
    else:
        margin_price = apply_margin(master.price)

    # 옵션 상품(option_list_price 있음): 정가 - 즉시할인 = 실효가 (sale_price 시점과 무관하게 일관됨)
    # 일반 즉시할인 상품(option_list_price 없음): sale_price(정가) - 즉시할인 = 실효가
    if store.option_list_price:
        effective_price = store.option_list_price - (store.option_discount_amount or 0)
    else:
        effective_price = store.sale_price - (store.option_discount_amount or 0)

    # 50원 이하 차이는 반올림·즉시할인 등 노이즈로 보고 신호 생성 생략
    if abs(margin_price - effective_price) <= 50:
        return

    # 옵션 구조 불일치 시 PRICE 신호 생성 생략 — OPTION_ADD / OPTION_PRICE_CHANGE 에 위임
    if master.options_text:
        if not master.option_diffs:
            return  # 패턴 A: master.options_text 있으나 option_diffs 없음 (데이터 이상치)
        if not store.option_list_price:
            return  # 패턴 B: master는 옵션인데 store는 단품 구조 (옵션 재구성 필요)
        if store.applied_option_diffs and master.option_diffs != store.applied_option_diffs:
            return  # 패턴 D: 옵션 차액 불일치 (OPTION_PRICE_CHANGE 대상)

    if margin_price > effective_price:
        if (master.id, store.id, "PRICE_UP_NEEDED") not in pending:
            db.session.add(ActionSignal(
                master_product_id=master.id,
                store_product_id=store.id,
                signal_type="PRICE_UP_NEEDED",
                current_value=json.dumps({"sale_price": effective_price}),
                suggested_value=json.dumps({"sale_price": master.price}),  # 도매가 저장, 실행 시 마진 재적용
            ))
            pending.add((master.id, store.id, "PRICE_UP_NEEDED"))
            stats["PRICE_UP_NEEDED"] += 1

    elif margin_price < effective_price:
        if (master.id, store.id, "PRICE_DOWN_POSSIBLE") not in pending:
            db.session.add(ActionSignal(
                master_product_id=master.id,
                store_product_id=store.id,
                signal_type="PRICE_DOWN_POSSIBLE",
                current_value=json.dumps({"sale_price": effective_price}),
                suggested_value=json.dumps({"sale_price": master.price}),  # 도매가 저장, 실행 시 마진 재적용
            ))
            pending.add((master.id, store.id, "PRICE_DOWN_POSSIBLE"))
            stats["PRICE_DOWN_POSSIBLE"] += 1


def _is_dimension_mismatch_blocked(store: StoreProduct) -> bool:
    """옵션 차원 불일치 (master 1차원 vs Naver 다차원) pending 상태인 상품인지.
    True면 OPTION_ADD / OPTION_STOCK_CHANGE / OPTION_PRICE_CHANGE 자동 시그널 생성 금지.
    기존 pending 시그널은 보존 (자동 정리 안 함)."""
    mm = getattr(store, "option_mismatch", None)
    return bool(mm and mm.status == "pending" and mm.mismatch_type == "dimension_mismatch")


def _check_option_signals(master: MasterProduct, store: StoreProduct, stats: dict, pending: set, prev_opts: dict):
    key = (master.id, store.id, "OPTION_PRICE_CHANGE")
    existing = prev_opts.get(key)

    # 옵션 차원 불일치 가드: 자동 시그널 생성 차단 (기존 pending은 보존)
    if _is_dimension_mismatch_blocked(store):
        return

    # 단품(옵션명 없음) → 옵션 가격 변동 비교 자체가 무의미
    if not master.options_text:
        if existing:
            db.session.delete(existing)
            prev_opts.pop(key, None)
            pending.discard(key)
        return
    if not master.price or not store.origin_product_no:
        return

    if (master.id, store.id, "PRICE_UP_NEEDED") in pending or (master.id, store.id, "PRICE_DOWN_POSSIBLE") in pending:
        return
    if (master.id, store.id, "OPTION_ADD") in pending:
        return

    # 1순위: StoreProduct에 저장된 적용 이력 확인
    # applied 쪽이 "0\n0\n0" 같은 전부-0 문자열이면 NULL 과 동일 취급 (master 는 이미 정규화됨)
    if _normalize_diffs_for_compare(store.applied_option_diffs) == master.option_diffs:
        if existing:
            db.session.delete(existing)
            prev_opts.pop(key, None)
            pending.discard(key)
        return

    last_executed = (
        ActionSignal.query
        .filter_by(store_product_id=store.id, signal_type="OPTION_PRICE_CHANGE")
        .filter(ActionSignal.status.in_(["executed", "reverted"]))
        .order_by(ActionSignal.resolved_at.desc())
        .first()
    )
    if last_executed:
        last_suggested = json.loads(last_executed.suggested_value or "{}")
        if _normalize_diffs_for_compare(last_suggested.get("option_diffs")) == master.option_diffs:
            if existing:
                db.session.delete(existing)
                prev_opts.pop(key, None)
                pending.discard(key)
            return

    # OPTION_ADD 실행 이력도 확인 — OPTION_ADD 실행 시 동일 option_diffs 적용됨
    last_add_exec = (
        ActionSignal.query
        .filter_by(store_product_id=store.id, signal_type="OPTION_ADD")
        .filter(ActionSignal.status.in_(["executed", "reverted"]))
        .order_by(ActionSignal.resolved_at.desc())
        .first()
    )
    if last_add_exec:
        last_add = json.loads(last_add_exec.suggested_value or "{}")
        if _normalize_diffs_for_compare(last_add.get("option_diffs")) == master.option_diffs:
            if existing:
                db.session.delete(existing)
                prev_opts.pop(key, None)
                pending.discard(key)
            return

    if key not in pending:
        from app.settings import calculate_option_pricing
        # master.option_diffs 가 NULL(추가금 전부 0)이어도 pricing 계산 가능하도록 옵션 개수만큼 0 채움
        # 실행 경로(OPTION_PRICE_CHANGE 1072 라인)도 빈 diffs 대신 "0\n0\n0" 형태를 받아야 정상 동작
        n_opts = len([x for x in master.options_text.split("\n") if x.strip()])
        diffs_for_calc = master.option_diffs or ("\n".join(["0"] * n_opts) if n_opts else "")
        pricing = calculate_option_pricing(master.price, diffs_for_calc)
        new_suggested = json.dumps({
            "base_price":   master.price,
            "option_diffs": diffs_for_calc,
            "options_text": master.options_text,
            "list_price":   pricing["list_price"],
            "discount":     pricing["discount"],
            "sale_price":   pricing["sale_price"],
            "additions":    pricing["additions"],
        })
        if existing:
            old = json.loads(existing.suggested_value or "{}")
            # 기존 pending 의 diffs 도 정규화해서 비교 — 같은 detect 를 여러 번 돌려도 중복 업데이트 방지
            if _normalize_diffs_for_compare(old.get("option_diffs")) == master.option_diffs:
                return  # 값 동일, 기존 pending 유지
            existing.suggested_value = new_suggested
            existing.detected_at = kst_now()
            stats["OPTION_PRICE_CHANGE"] += 1
        else:
            db.session.add(ActionSignal(
                master_product_id=master.id,
                store_product_id=store.id,
                signal_type="OPTION_PRICE_CHANGE",
                current_value=json.dumps({
                    "options_text": store.applied_options_text or "",
                    "store_list_price": store.option_list_price,
                    "store_discount": store.option_discount_amount,
                }),
                suggested_value=new_suggested,
            ))
            pending.add(key)
            stats["OPTION_PRICE_CHANGE"] += 1


def _check_option_stock_signals(master: MasterProduct, store: StoreProduct, stats: dict, pending: set, prev_opts: dict):
    """옵션 재고 변동 감지 — 수집기가 extra["옵션재고"] 제공 시 동작"""
    key = (master.id, store.id, "OPTION_STOCK_CHANGE")
    existing = prev_opts.get(key)

    # 옵션 차원 불일치 가드: 자동 시그널 생성 차단 (기존 pending은 보존)
    if _is_dimension_mismatch_blocked(store):
        return

    if not master.options_text or master.option_stocks is None:
        if existing:
            db.session.delete(existing)
            prev_opts.pop(key, None)
            pending.discard(key)
        return
    if not store.origin_product_no:
        return

    last_executed = (
        ActionSignal.query
        .filter_by(store_product_id=store.id, signal_type="OPTION_STOCK_CHANGE")
        .filter(ActionSignal.status.in_(["executed", "reverted"]))
        .order_by(ActionSignal.resolved_at.desc())
        .first()
    )
    if last_executed:
        last_suggested = json.loads(last_executed.suggested_value or "{}")
        if last_suggested.get("option_stocks") == master.option_stocks:
            if existing:
                db.session.delete(existing)
                prev_opts.pop(key, None)
                pending.discard(key)
            return

    if key not in pending:
        new_suggested = json.dumps({
            "option_stocks": master.option_stocks,
            "options_text":  master.options_text,
        })
        if existing:
            old = json.loads(existing.suggested_value or "{}")
            if old.get("option_stocks") == master.option_stocks:
                return  # 값 동일, 기존 pending 유지
            existing.suggested_value = new_suggested
            existing.detected_at = kst_now()
            stats["OPTION_STOCK_CHANGE"] += 1
        else:
            db.session.add(ActionSignal(
                master_product_id=master.id,
                store_product_id=store.id,
                signal_type="OPTION_STOCK_CHANGE",
                current_value=json.dumps({"option_stocks": master.option_stocks}),
                suggested_value=new_suggested,
            ))
            pending.add(key)
            stats["OPTION_STOCK_CHANGE"] += 1


def _check_option_stock_refill_signals(master: MasterProduct, store: StoreProduct, stats: dict, pending: set, prev_opts: dict):
    """SALE 상태 + 도매처 active 인데 Naver 옵션 재고만 0인 옵션 감지.
    sync_store_option_state 가 캐시한 stockQuantity / usable 기반.
    가드:
      - 캐시 데이터 없음 → 스킵 (sync 미실행)
      - master.current_status != active → 스킵
      - 같은 상품에 OPTION_STOCK_CHANGE pending → 스킵 (master 기준 보정이 우선)
      - master.option_stocks 가 있고 0 포함 → 스킵 (도매처도 품절이라 보정 위험)
    """
    key = (master.id, store.id, "OPTION_STOCK_REFILL_NEEDED")
    existing = prev_opts.get(key)

    def _drop():
        if existing:
            db.session.delete(existing)
            prev_opts.pop(key, None)
            pending.discard(key)

    if not store.naver_cached_option_stocks or not store.naver_cached_option_usable:
        _drop()
        return
    if master.current_status != "active":
        _drop()
        return
    if not store.origin_product_no:
        _drop()
        return
    if (master.id, store.id, "OPTION_STOCK_CHANGE") in pending:
        _drop()
        return
    if master.option_stocks:
        try:
            if any(int(v.strip()) == 0 for v in master.option_stocks.split("\n") if v.strip()):
                _drop()
                return
        except ValueError:
            pass

    try:
        stocks = [s.strip() for s in store.naver_cached_option_stocks.split("\n") if s.strip()]
        usables = [u.strip() for u in store.naver_cached_option_usable.split("\n") if u.strip()]
    except Exception:
        return
    if len(stocks) != len(usables) or not stocks:
        return

    refill_indices = [i for i, (stk, us) in enumerate(zip(stocks, usables)) if us == "1" and stk == "0"]
    if not refill_indices:
        _drop()
        return

    new_current = json.dumps({"option_stocks_cache": store.naver_cached_option_stocks})
    new_suggested = json.dumps({"refill_count": len(refill_indices), "refill_indices": refill_indices})

    if existing:
        old_sv = json.loads(existing.suggested_value or "{}")
        if old_sv.get("refill_indices") == refill_indices:
            return  # 동일 — 기존 pending 유지
        existing.current_value = new_current
        existing.suggested_value = new_suggested
        existing.detected_at = kst_now()
        stats["OPTION_STOCK_REFILL_NEEDED"] += 1
    else:
        db.session.add(ActionSignal(
            master_product_id=master.id,
            store_product_id=store.id,
            signal_type="OPTION_STOCK_REFILL_NEEDED",
            current_value=new_current,
            suggested_value=new_suggested,
        ))
        pending.add(key)
        stats["OPTION_STOCK_REFILL_NEEDED"] += 1


def _check_option_add_signals(master: MasterProduct, store: StoreProduct, stats: dict, pending: set, prev_opts: dict):
    """도매처에 새 옵션이 추가되었거나 옵션 구성이 바뀐 경우 감지"""
    key = (master.id, store.id, "OPTION_ADD")
    existing = prev_opts.get(key)

    # 옵션 차원 불일치 가드: 자동 시그널 생성 차단 (기존 pending은 보존)
    if _is_dimension_mismatch_blocked(store):
        return

    if not master.options_text:
        if existing:
            db.session.delete(existing)
            prev_opts.pop(key, None)
            pending.discard(key)
        return
    # 추가금 없는 상품 + applied 이력 없음: 실행 이력 확인 후 판단 (flooding 방지)
    if not _has_extra_price(master) and store.applied_options_text is None:
        # naver_cached_additions 기반 사전 보정:
        # Naver 실측 추가금이 모두 0 이고 옵션 개수가 master 와 일치하면
        # = Naver 가 이미 동일 옵션 구조를 +0 으로 보유 → 신호 생성 보류 + applied_* 동기화
        if store.naver_cached_additions:
            cached_lines = [l.strip() for l in store.naver_cached_additions.split("\n") if l.strip()]
            master_lines = [l.strip() for l in master.options_text.split("\n") if l.strip()]
            if len(cached_lines) == len(master_lines):
                try:
                    if all(int(x) == 0 for x in cached_lines):
                        store.applied_options_text = master.options_text
                        store.applied_option_diffs = None
                        store.applied_option_base_price = master.price
                        if existing:
                            db.session.delete(existing)
                            prev_opts.pop(key, None)
                            pending.discard(key)
                        return
                except ValueError:
                    pass

        last_add = (
            ActionSignal.query
            .filter_by(store_product_id=store.id, signal_type="OPTION_ADD")
            .filter(ActionSignal.status == "executed")
            .order_by(ActionSignal.resolved_at.desc())
            .first()
        )
        if last_add:
            last_sv = json.loads(last_add.suggested_value or "{}")
            if last_sv.get("options_text") == master.options_text:
                # 이전에 동일 옵션으로 실행됨 → applied 이력 동기화 후 스킵
                store.applied_options_text = master.options_text
                if existing:
                    db.session.delete(existing)
                    prev_opts.pop(key, None)
                    pending.discard(key)
                return
        # 실행 이력 없음 → Naver에 옵션이 없을 수 있으므로 OPTION_ADD 신호 생성
    if not master.price or not store.origin_product_no:
        return

    # 1순위: StoreProduct에 저장된 적용 이력 확인 (ActionSignal 이력 없어도 동작)
    if (store.applied_options_text == master.options_text and
            store.applied_option_diffs == master.option_diffs and
            store.applied_option_base_price == master.price):
        if existing:
            db.session.delete(existing)
            prev_opts.pop(key, None)
            pending.discard(key)
        return

    last_executed = (
        ActionSignal.query
        .filter_by(store_product_id=store.id, signal_type="OPTION_ADD")
        .filter(ActionSignal.status.in_(["executed", "reverted"]))
        .order_by(ActionSignal.resolved_at.desc())
        .first()
    )
    if last_executed:
        last_suggested = json.loads(last_executed.suggested_value or "{}")
        if last_suggested.get("options_text") == master.options_text:
            if existing:
                db.session.delete(existing)
                prev_opts.pop(key, None)
                pending.discard(key)
            return

    last_price_exec = (
        ActionSignal.query
        .filter_by(store_product_id=store.id, signal_type="OPTION_PRICE_CHANGE")
        .filter(ActionSignal.status.in_(["executed", "reverted"]))
        .order_by(ActionSignal.resolved_at.desc())
        .first()
    )
    if last_price_exec:
        last_p = json.loads(last_price_exec.suggested_value or "{}")
        if last_p.get("options_text") == master.options_text:
            if existing:
                db.session.delete(existing)
                prev_opts.pop(key, None)
                pending.discard(key)
            return

    new_suggested = json.dumps({
        "base_price":   master.price,
        "option_diffs": master.option_diffs,
        "options_text": master.options_text,
    })
    new_current = json.dumps({"options_text": store.applied_options_text or ""})
    if existing:
        old = json.loads(existing.suggested_value or "{}")
        if (old.get("options_text") == master.options_text and
                old.get("option_diffs") == master.option_diffs):
            return  # 값 동일, 기존 pending 유지
        existing.current_value = new_current
        existing.suggested_value = new_suggested
        existing.detected_at = kst_now()
        stats["OPTION_ADD"] += 1
    elif key not in pending:
        db.session.add(ActionSignal(
            master_product_id=master.id,
            store_product_id=store.id,
            signal_type="OPTION_ADD",
            current_value=new_current,
            suggested_value=new_suggested,
        ))
        pending.add(key)
        stats["OPTION_ADD"] += 1


def _check_status_signals(master: MasterProduct, store: StoreProduct, stats: dict, pending: set):
    store_active = store.store_status == "SALE"
    master_status = master.current_status

    if master_status == "discontinued":
        if store_active and (master.id, store.id, "DISCONTINUE_NEEDED") not in pending:
            db.session.add(ActionSignal(
                master_product_id=master.id,
                store_product_id=store.id,
                signal_type="DISCONTINUE_NEEDED",
                current_value=json.dumps({"store_status": store.store_status}),
                suggested_value=json.dumps({"store_status": "CLOSE"}),
            ))
            pending.add((master.id, store.id, "DISCONTINUE_NEEDED"))
            stats["DISCONTINUE_NEEDED"] += 1

    elif master_status in ("missing", "discontinued_candidate"):
        if store_active and (master.id, store.id, "SUSPEND_NEEDED") not in pending:
            db.session.add(ActionSignal(
                master_product_id=master.id,
                store_product_id=store.id,
                signal_type="SUSPEND_NEEDED",
                current_value=json.dumps({"store_status": store.store_status}),
                suggested_value=json.dumps({"store_status": "SUSPENSION"}),
            ))
            pending.add((master.id, store.id, "SUSPEND_NEEDED"))
            stats["SUSPEND_NEEDED"] += 1

    elif master_status == "out_of_stock":
        # 도매처에서 품절 → 상품 전체 중지 (옵션 유무 무관)
        if store_active and (master.id, store.id, "SUSPEND_NEEDED") not in pending:
            db.session.add(ActionSignal(
                master_product_id=master.id,
                store_product_id=store.id,
                signal_type="SUSPEND_NEEDED",
                current_value=json.dumps({"store_status": store.store_status}),
                suggested_value=json.dumps({"store_status": "SUSPENSION"}),
            ))
            pending.add((master.id, store.id, "SUSPEND_NEEDED"))
            stats["SUSPEND_NEEDED"] += 1

    elif master_status == "active":
        # CLOSE(판매종료)는 API로 복구 불가 → 제외
        resumable = not store_active and store.store_status != "CLOSE"
        if resumable and (master.id, store.id, "RESUME_POSSIBLE") not in pending:
            db.session.add(ActionSignal(
                master_product_id=master.id,
                store_product_id=store.id,
                signal_type="RESUME_POSSIBLE",
                current_value=json.dumps({"store_status": store.store_status}),
                suggested_value=json.dumps({"store_status": "SALE"}),
            ))
            pending.add((master.id, store.id, "RESUME_POSSIBLE"))
            stats["RESUME_POSSIBLE"] += 1
