"""
스마트스토어 일괄등록 엑셀 생성기
- 93컬럼 구조 (A~CO)
- 1행: 카테고리 그룹 헤더
- 2행: 컬럼명 헤더
- 3행~: 상품 데이터
"""
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# 93개 컬럼 헤더 (인덱스 0부터)
HEADERS = [
    "판매자 상품코드",     # A  0
    "카테고리코드",        # B  1  ← 빈칸 (사용자 입력)
    "상품명",             # C  2
    "상품상태",           # D  3
    "판매가",             # E  4
    "부가세",             # F  5
    "재고수량(옵션없을경우)",  # G  6
    "모델번호",           # H  7
    "브랜드",             # I  8
    "제조사",             # J  9
    "원산지코드",         # K  10
    "원산지 직접입력",     # L  11
    "상품무게(kg)",       # M  12
    "발송가능일",         # N  13
    "대표이미지",         # O  14
    "추가이미지1",        # P  15
    "추가이미지2",        # Q  16
    "추가이미지3",        # R  17
    "추가이미지4",        # S  18
    "추가이미지5",        # T  19
    "추가이미지6",        # U  20
    "추가이미지7",        # V  21
    "추가이미지8",        # W  22
    "추가이미지9",        # X  23
    "상세설명",           # Y  24
    "PC검색어1",          # Z  25
    "PC검색어2",          # AA 26
    "PC검색어3",          # AB 27
    "PC검색어4",          # AC 28
    "PC검색어5",          # AD 29
    "모바일검색어1",       # AE 30
    "모바일검색어2",       # AF 31
    "모바일검색어3",       # AG 32
    "모바일검색어4",       # AH 33
    "모바일검색어5",       # AI 34
    "배송방법",           # AJ 35
    "직접입력 택배사",     # AK 36
    "배송비유형",         # AL 37
    "기본배송비",         # AM 38
    "배송비 결제방식",     # AN 39
    "조건부무료-상품판매가합계",  # AO 40
    "제주/도서지방 추가배송비",  # AP 41
    "설치비",            # AQ 42
    "일괄배송",          # AR 43
    "묶음배송여부",       # AS 44
    "발송기준일",         # AT 45
    "반품배송비",         # AU 46
    "교환배송비",         # AV 47
    "반품/교환 택배사",   # AW 48
    "반품지(상품반송지) 명칭",  # AX 49
    "반품지 우편번호",    # AY 50
    "반품지 기본주소",    # AZ 51
    "반품지 상세주소",    # BA 52
    "출고지 명칭",        # BB 53
    "출고지 우편번호",    # BC 54
    "출고지 기본주소",    # BD 55
    "출고지 상세주소",    # BE 56
    "A/S 전화번호",       # BF 57
    "A/S 안내",          # BG 58
    "고시정보1",         # BH 59
    "고시정보2",         # BI 60
    "고시정보3",         # BJ 61
    "고시정보4",         # BK 62
    "고시정보5",         # BL 63
    "고시정보6",         # BM 64
    "고시정보7",         # BN 65
    "고시정보8",         # BO 66
    "고시정보9",         # BP 67
    "고시정보10",        # BQ 68
    "고시정보11",        # BR 69
    "고시정보12",        # BS 70
    "고시정보13",        # BT 71
    "고시정보14",        # BU 72
    "고시정보15",        # BV 73
    "고시정보16",        # BW 74
    "고시정보17",        # BX 75
    "고시정보18",        # BY 76
    "고시정보19",        # BZ 77
    "고시정보20",        # CA 78
    "고시정보21",        # CB 79
    "고시정보22",        # CC 80
    "고시정보23",        # CD 81
    "고시정보24",        # CE 82
    "고시정보25",        # CF 83
    "미성년자구매불가여부",  # CG 84
    "구매수량제한",       # CH 85
    "전시여부",          # CI 86
    "판매여부",          # CJ 87
    "판매시작일",        # CK 88
    "판매종료일",        # CL 89
    "구매평적립금",      # CM 90
    "리뷰적립금",        # CN 91
    "즉시할인가",        # CO 92
]

# 자동 입력 컬럼 매핑 (0-based 인덱스 → callable or 고정값)
FILL_MAP = {
    0:  lambda p: p.supplier_product_code,          # A: 판매자 상품코드
    # 1 (B): 카테고리코드 — 빈칸
    2:  lambda p: p.product_name,                   # C: 상품명
    3:  "신상품",                                    # D: 상품상태
    4:  lambda p: p.price,                          # E: 판매가
    5:  "과세상품",                                  # F: 부가세
    6:  999,                                         # G: 재고수량
    10: "0200",                                      # K: 원산지코드 (국내산)
    14: lambda p: p.image_url or "",                # O: 대표이미지
    24: lambda p: f'<img src="{p.image_url}">' if p.image_url else "",  # Y: 상세설명
    35: "택배,소포,등기",                            # AJ: 배송방법
    37: "조건부무료",                                # AL: 배송비유형
    38: 3000,                                        # AM: 기본배송비
    39: "착불또는선결제",                            # AN: 배송비 결제방식
    40: 30000,                                       # AO: 조건부무료 조건금액
    46: 2500,                                        # AU: 반품배송비
    47: 2500,                                        # AV: 교환배송비
}


def _cell_value(mapping_val, product):
    if callable(mapping_val):
        return mapping_val(product)
    return mapping_val


def generate_smartstore_excel(wholesaler_id: int, status_filter: str = "active") -> bytes:
    from app.master.models import MasterProduct

    if status_filter == "all":
        products = (
            MasterProduct.query
            .filter_by(wholesaler_id=wholesaler_id)
            .order_by(MasterProduct.supplier_product_code)
            .all()
        )
    else:
        products = (
            MasterProduct.query
            .filter_by(wholesaler_id=wholesaler_id, current_status=status_filter)
            .order_by(MasterProduct.supplier_product_code)
            .all()
        )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "일괄등록"

    # 1행: 그룹 헤더 (스마트스토어 양식 형식)
    group_labels = {
        0:  "상품 기본정보",
        14: "이미지",
        24: "상세설명",
        25: "검색어",
        35: "배송정보",
        49: "반품/교환정보",
        57: "A/S정보",
        59: "고시정보",
        84: "판매정보",
    }
    for col_idx, label in group_labels.items():
        cell = ws.cell(row=1, column=col_idx + 1, value=label)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2C3E50")
        cell.alignment = Alignment(horizontal="center")

    # 2행: 컬럼명 헤더
    header_fill = PatternFill("solid", fgColor="ECF0F1")
    for col_idx, header in enumerate(HEADERS):
        cell = ws.cell(row=2, column=col_idx + 1, value=header)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # 3행~: 상품 데이터
    for row_idx, product in enumerate(products, start=3):
        for col_idx in range(93):
            val = _cell_value(FILL_MAP.get(col_idx), product) if col_idx in FILL_MAP else None
            ws.cell(row=row_idx, column=col_idx + 1, value=val)

    # 컬럼 너비 조정
    ws.column_dimensions["A"].width = 18  # 판매자 상품코드
    ws.column_dimensions["B"].width = 14  # 카테고리코드
    ws.column_dimensions["C"].width = 40  # 상품명
    ws.column_dimensions["E"].width = 10  # 판매가
    ws.column_dimensions["O"].width = 50  # 대표이미지
    ws.column_dimensions["Y"].width = 30  # 상세설명

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
